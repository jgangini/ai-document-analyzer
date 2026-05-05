from __future__ import annotations

import csv
from datetime import date
import json
from pathlib import Path
from types import MethodType

import pytest

from apps.backend.app.api.routes.questions import (
    _build_citations_and_sources,
    _extract_latest_conversation_scope_from_messages,
    _load_visible_scope_options,
)
from apps.backend.app.api.routes.metadata import _map_upload_summary, _parse_row_preview
from apps.backend.app.api.contracts.questions import EvidenceItem
from apps.backend.app.core.config import Settings
from apps.backend.app.agent.contracts import LLMResult
from apps.backend.app.agent.nodes import QAGraphNodes, _build_retrieval_question
from apps.backend.app.agent.router import GraphSynthesis
from apps.backend.app.agent.tools.hybrid_answer_tool import HybridAnswerTool
from apps.backend.app.agent.tools.multimodal_tool import VisualInspectionResult
from apps.backend.app.ingest.document_ingest_service import IngestionService
from apps.backend.app.ingest.document_metadata import FileMetadata, extract_document_code_from_filename
from apps.backend.app.ingest.rag_enrichment import _extract_secondary_identifier, build_file_group_key
from apps.backend.app.rag.embedding_service import EmbeddingService, NomicLocalMultimodalProvider
from apps.backend.app.rag.facts_query_service import ArchiveMetadataEntry, FactResolution, QuestionFactResolver
from apps.backend.app.rag.display_text import repair_document_file_name
from apps.backend.app.rag.query_selectors import (
    ParsedQuestionSelectors,
    build_effective_selector_question,
    merge_question_selectors,
    parse_question_selectors,
)
from apps.backend.app.rag.question_classifier import QuestionClassifier
from apps.backend.app.rag.retrieval.query_service import (
    RetrievalResult,
    RetrievalPipelineService,
    extract_explicit_file_references,
    question_requests_full_document_coverage,
    question_requires_visual_grounding,
    question_requests_representative_details,
)
from apps.backend.app.rag.retrieval.oracle_vector_search import OracleVectorSearchResult
from apps.backend.app.repositories.archive_metadata_repository import ArchiveMetadataRepository
from apps.backend.app.repositories.repository_utils import build_oracle_text_contains_query
from apps.backend.app.rag.scope_resolver import (
    QuestionScopeResolver,
    ScopeResolutionError,
    extract_candidate_archive_slugs_from_question,
    extract_candidate_codes_from_question,
    extract_candidate_file_names_from_question,
)
from apps.backend.app.services.metadata_upload_service import (
    MetadataUploadService,
    MetadataUploadValidationError,
)
from apps.backend.app.services.metadata_normalization_service import (
    LoadedWorkbookSheet,
    MetadataWorkbookNormalizationError,
    normalize_metadata_workbook_to_csv,
)


class _FakeScopeRepository:
    def __init__(self) -> None:
        self._files_by_user = {
            7: [101, 102, 103, 104],
        }
        self._codes_by_user = {
            7: {
                "AI041": [101, 102],
                "RM797": [101, 102],
                "RM798": [103, 104],
            }
        }
        self._archive_slugs_by_user = {
            7: {
                "RM797_ID_1668": [101],
                "RM797_ID_5515": [102],
            }
        }
        self._file_names_by_user = {
            7: {
                "ai041.pdf": [101],
                "ai041_modificacion.pdf": [102],
            }
        }

    def filter_file_ids_for_user(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> list[int]:
        del include_shared
        allowed = set(self._files_by_user.get(int(user_id), []))
        return [file_id for file_id in file_ids if int(file_id) in allowed]

    def list_distinct_document_codes_for_user(
        self,
        *,
        user_id: int,
        include_shared: bool = False,
    ) -> list[str]:
        del include_shared
        return sorted(self._codes_by_user.get(int(user_id), {}).keys())

    def list_file_ids_for_document_codes(
        self,
        *,
        user_id: int,
        document_codes: list[str],
        include_shared: bool = False,
    ) -> list[int]:
        del include_shared
        resolved: list[int] = []
        for code in document_codes:
            resolved.extend(self._codes_by_user.get(int(user_id), {}).get(code, []))
        return resolved

    def list_known_archive_slugs_for_user(
        self,
        *,
        user_id: int,
        include_shared: bool = False,
    ) -> list[str]:
        del include_shared
        return sorted(self._archive_slugs_by_user.get(int(user_id), {}).keys())

    def list_file_ids_for_archive_slugs(
        self,
        *,
        user_id: int,
        archive_slugs: list[str],
        include_shared: bool = False,
    ) -> list[int]:
        del include_shared
        resolved: list[int] = []
        for archive_slug in archive_slugs:
            resolved.extend(self._archive_slugs_by_user.get(int(user_id), {}).get(archive_slug, []))
        return resolved

    def list_file_ids_for_input_filenames(
        self,
        *,
        user_id: int,
        file_names: list[str],
        file_ids: list[int] | None = None,
        include_shared: bool = False,
    ) -> list[int]:
        del include_shared
        allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
        resolved: list[int] = []
        for file_name in file_names:
            for file_id in self._file_names_by_user.get(int(user_id), {}).get(str(file_name).lower(), []):
                if allowed and int(file_id) not in allowed:
                    continue
                resolved.append(int(file_id))
        return resolved

    def get_archive_slug_map_for_file_ids(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> dict[int, str]:
        del include_shared
        allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
        resolved: dict[int, str] = {}
        for archive_slug, scoped_file_ids in self._archive_slugs_by_user.get(int(user_id), {}).items():
            for file_id in scoped_file_ids:
                if int(file_id) in allowed:
                    resolved[int(file_id)] = str(archive_slug)
        return resolved

    def count_files_for_user(self, *, user_id: int, include_shared: bool = False) -> int:
        del include_shared
        return len(self._files_by_user.get(int(user_id), []))


class _FakeMetadataRepository:
    def __init__(self) -> None:
        self.created_uploads: list[dict[str, object]] = []
        self.replaced_rows: list[dict[str, object]] = []
        self.updated_uploads: list[dict[str, object]] = []
        self.refresh_calls: list[dict[str, object]] = []

    def list_known_archive_slugs_for_user(self, *, user_id: int, include_shared: bool = False) -> list[str]:
        del user_id
        del include_shared
        return ["RM797_ID_1668", "RM797_ID_5515"]

    def create_upload(
        self,
        *,
        metadata_upload_id: str,
        user_id: int,
        source_file_name: str,
        columns: list[str],
        total_rows: int,
        display_name: str | None = None,
        description: str | None = None,
        access_scope: str | None = None,
    ) -> dict[str, object]:
        payload = {
            "metadata_upload_id": metadata_upload_id,
            "user_id": user_id,
            "source_file_name": source_file_name,
            "display_name": display_name or source_file_name,
            "description": description or "",
            "access_scope": access_scope or "private",
            "metadata_status": "active",
            "column_names_json": columns,
            "total_rows": total_rows,
        }
        self.created_uploads.append(payload)
        return {
            **payload,
            "metadata_upload_created": date(2026, 4, 21),
        }

    def replace_upload_rows(
        self,
        *,
        metadata_upload_id: str,
        user_id: int,
        rows: list[dict[str, object]],
    ) -> None:
        self.replaced_rows.append(
            {
                "metadata_upload_id": metadata_upload_id,
                "user_id": user_id,
                "rows": rows,
            }
        )

    def update_upload_content(
        self,
        *,
        metadata_upload_id: str,
        user_id: int,
        source_file_name: str,
        columns: list[str],
        total_rows: int,
    ) -> dict[str, object] | None:
        payload = {
            "metadata_upload_id": metadata_upload_id,
            "user_id": user_id,
            "source_file_name": source_file_name,
            "display_name": source_file_name,
            "description": "",
            "metadata_status": "active",
            "column_names_json": columns,
            "total_rows": total_rows,
            "metadata_upload_created": date(2026, 4, 21),
        }
        self.updated_uploads.append(payload)
        return payload

    def refresh_linked_archive_metadata_from_upload(self, *, metadata_upload_id: str, user_id: int) -> int:
        self.refresh_calls.append({"metadata_upload_id": metadata_upload_id, "user_id": user_id})
        return 0


class _FakeArchiveMetadataFileRepository:
    def __init__(
        self,
        rows: list[dict[str, object]],
        *,
        file_rows: list[dict[str, object]] | None = None,
        page_quality_rows: list[dict[str, object]] | None = None,
    ) -> None:
        self.rows = [dict(row) for row in rows]
        self.file_rows = [dict(row) for row in list(file_rows or [])]
        self.page_quality_rows = [dict(row) for row in list(page_quality_rows or [])]

    def get_archive_metadata_for_file_ids(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> list[dict[str, object]]:
        del user_id, include_shared
        allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
        return [
            dict(row)
            for row in self.rows
            if int(row.get("file_id") or 0) in allowed
        ]

    def list_archive_metadata_for_user(self, *, user_id: int, include_shared: bool = False) -> list[dict[str, object]]:
        del user_id, include_shared
        return [dict(row) for row in self.rows]

    def get_archive_slug_map_for_file_ids(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> dict[int, str]:
        del user_id, include_shared
        allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
        return {
            int(row.get("file_id") or 0): str(row.get("archive_slug") or "")
            for row in self.rows
            if int(row.get("file_id") or 0) in allowed and str(row.get("archive_slug") or "").strip()
        }

    def list_files_for_user(self, *, user_id: int, include_shared: bool = False) -> list[dict[str, object]]:
        del user_id, include_shared
        return [dict(row) for row in self.file_rows]

    def list_page_quality_for_file_ids(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> list[dict[str, object]]:
        del user_id, include_shared
        allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
        return [
            dict(row)
            for row in self.page_quality_rows
            if int(row.get("file_id") or 0) in allowed
        ]


def test_scope_options_preserve_metadata_upload_csv_column_order() -> None:
    repository = _FakeArchiveMetadataFileRepository(
        [
            {
                "file_id": 101,
                "archive_slug": "AI041_ID_49",
                "column_names_json": json.dumps(
                    [
                        "file",
                        "Id",
                        "Usuario",
                        "Codigo de Sitio",
                        "Nombre de Sitio",
                        "Clasificación de Sitio",
                        "Monto",
                    ],
                    ensure_ascii=False,
                ),
                "metadata_json": json.dumps(
                    {
                        "file": "AI041_ID_49",
                        "fields": {
                            "Monto": 123,
                            "Id": 49,
                        },
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        file_rows=[
            {
                "file_id": 101,
                "archive_slug": "AI041_ID_49",
            }
        ],
    )

    options = _load_visible_scope_options(repository=repository, user_id=7)

    assert options.metadata_fields == [
        "Id",
        "Usuario",
        "Codigo de Sitio",
        "Nombre de Sitio",
        "Clasificación de Sitio",
        "Monto",
    ]


class _FakeInventoryFileRepository:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self.rows = [dict(row) for row in rows]

    def list_files_for_user(self, *, user_id: int, include_shared: bool = False) -> list[dict[str, object]]:
        del user_id, include_shared
        return [dict(row) for row in self.rows]


class _RecordingScopeRepository(_FakeScopeRepository):
    def __init__(self) -> None:
        super().__init__()
        self.calls: list[tuple[str, bool]] = []

    def filter_file_ids_for_user(
        self,
        *,
        user_id: int,
        file_ids: list[int],
        include_shared: bool = False,
    ) -> list[int]:
        self.calls.append(("filter_file_ids_for_user", bool(include_shared)))
        return super().filter_file_ids_for_user(
            user_id=user_id,
            file_ids=file_ids,
            include_shared=include_shared,
        )

    def list_distinct_document_codes_for_user(
        self,
        *,
        user_id: int,
        include_shared: bool = False,
    ) -> list[str]:
        self.calls.append(("list_distinct_document_codes_for_user", bool(include_shared)))
        return super().list_distinct_document_codes_for_user(
            user_id=user_id,
            include_shared=include_shared,
        )

    def list_known_archive_slugs_for_user(
        self,
        *,
        user_id: int,
        include_shared: bool = False,
    ) -> list[str]:
        self.calls.append(("list_known_archive_slugs_for_user", bool(include_shared)))
        return super().list_known_archive_slugs_for_user(
            user_id=user_id,
            include_shared=include_shared,
        )


class _RecordingInventoryFileRepository(_FakeInventoryFileRepository):
    def __init__(self, rows: list[dict[str, object]]) -> None:
        super().__init__(rows)
        self.include_shared_calls: list[bool] = []

    def list_files_for_user(self, *, user_id: int, include_shared: bool = False) -> list[dict[str, object]]:
        self.include_shared_calls.append(bool(include_shared))
        return super().list_files_for_user(user_id=user_id, include_shared=include_shared)


def _build_settings() -> Settings:
    return Settings(_env_file=None)


def _make_evidence_item(
    *,
    file_id: int,
    file_name: str,
    page_id: int | None = None,
    source_number: int | None = None,
    page_number: int = 1,
    summary_text: str | None = None,
    score: float = 0.9,
    ocr_confidence: float = 0.95,
) -> EvidenceItem:
    resolved_page_id = int(page_id or file_id)
    return EvidenceItem(
        source_number=int(source_number or resolved_page_id),
        file_id=int(file_id),
        file_name=file_name,
        file_code=None,
        page_id=resolved_page_id,
        page_number=int(page_number),
        score=float(score),
        summary_text=summary_text or f"Evidencia OCR relevante para {file_name}.",
        image_path_local="",
        object_name_page="",
        extraction_method="docling_rapidocr",
        ocr_confidence=float(ocr_confidence),
    )


def test_cited_sources_include_snippets_for_chat_page_preview() -> None:
    evidence = [
        _make_evidence_item(
            file_id=101,
            file_name="RM797-Contrato_2.pdf",
            source_number=1,
            page_number=8,
            summary_text="QUINTO: El precio de arrendamiento se pagara en las condiciones pactadas.",
        ),
        _make_evidence_item(
            file_id=102,
            file_name="RM797_Rectificacion.pdf",
            source_number=2,
            page_number=3,
            summary_text="Rectificacion de antecedentes generales del contrato.",
        ),
    ]

    citations, _sources, cited_sources, retrieved_sources = _build_citations_and_sources(
        analyzed_evidence=evidence,
        citation_numbers=[1],
    )

    assert citations[0].snippet.startswith("QUINTO")
    assert cited_sources[0].snippet.startswith("QUINTO")
    assert retrieved_sources[0].snippet.startswith("QUINTO")


def _build_ingestion_service_for_tests() -> IngestionService:
    return object.__new__(IngestionService)


def _build_metadata_upload_service(
    *,
    tmp_path: Path,
    repository: _FakeMetadataRepository | None = None,
) -> MetadataUploadService:
    uploads_dir = tmp_path / "uploads"
    extracted_dir = tmp_path / "extracted"
    staging_dir = tmp_path / "staging"
    uploads_dir.mkdir()
    extracted_dir.mkdir()
    staging_dir.mkdir()
    settings = Settings(
        _env_file=None,
        UPLOAD_DIR=str(uploads_dir),
        EXTRACTED_DIR=str(extracted_dir),
        STAGING_DIR=str(staging_dir),
    )
    return MetadataUploadService(
        settings=settings,
        repository=repository or _FakeMetadataRepository(),
    )


def _normalize_sql_whitespace(value: str) -> str:
    return " ".join(str(value or "").split())


def test_metadata_upload_summary_mapping_preserves_csv_column_order() -> None:
    summary = _map_upload_summary(
        {
            "metadata_upload_id": "upload-1",
            "source_file_name": "metadata.csv",
            "display_name": "Contracts metadata",
            "description": "Dynamic CSV",
            "metadata_status": "active",
            "column_names_json": json.dumps(["file", "Id", "Banco", "Monto"], ensure_ascii=False),
            "total_rows": 2,
            "row_count": 2,
            "matched_files_count": 1,
            "unmatched_files_count": 1,
            "linked_documents_count": 6,
            "metadata_upload_created": date(2026, 4, 21),
            "metadata_upload_updated": date(2026, 4, 22),
        }
    )

    assert summary.columns == ["file", "Id", "Banco", "Monto"]
    assert summary.display_name == "Contracts metadata"
    assert summary.matched_files_count == 1
    assert summary.linked_documents_count == 6


def test_metadata_row_preview_handles_dynamic_fields() -> None:
    row = _parse_row_preview(
        {
            "file_key": "LA122_ID_3979",
            "row_json": json.dumps(
                {
                    "file": "LA122_ID_3979",
                    "fields": {
                        "Campo Arbitrario": "valor",
                        "Monto": 1200,
                    },
                },
                ensure_ascii=False,
            ),
        }
    )

    assert row.file == "LA122_ID_3979"
    assert row.fields == {"Campo Arbitrario": "valor", "Monto": 1200}


def test_extract_document_code_from_filename() -> None:
    code, source = extract_document_code_from_filename("RM797_-_Decreto_MOP_Exento_N667.pdf")
    assert code == "RM797"
    assert source == "filename_rule"


def test_extract_document_code_from_filename_with_underscore_prefix() -> None:
    code, source = extract_document_code_from_filename("RM797_Contrato.pdf")
    assert code == "RM797"
    assert source == "filename_rule"


def test_extract_document_code_from_filename_with_dash_prefix() -> None:
    code, source = extract_document_code_from_filename("RM797-Contrato_2.pdf")
    assert code == "RM797"
    assert source == "filename_rule"


def test_extract_document_code_from_filename_without_separator() -> None:
    code, source = extract_document_code_from_filename("Decreto_MOP_Exento_N667.pdf")
    assert code is None
    assert source == "none"


def test_file_types_bootstrap_script_is_removed() -> None:
    sql_path = Path("apps/backend/db/bootstrap/sql/04_file_types.sql")
    assert sql_path.exists() is False


def test_archive_metadata_repository_is_bootstrap_only() -> None:
    assert hasattr(ArchiveMetadataRepository, "ensure_schema") is False


def test_extract_candidate_codes_from_question_ignores_single_letter_tokens() -> None:
    codes = extract_candidate_codes_from_question(
        "Genera un analisis de RM797 y RM798, pero no del decreto N667."
    )
    assert codes == ["RM797", "RM798"]


def test_extract_candidate_codes_from_question_ignores_pdf_basenames() -> None:
    codes = extract_candidate_codes_from_question(
        "Compara AI041_Carta_Aviso_Cesin_Contrato_Alba_ATC.pdf y "
        "ATC-Comunicacin_Entel_Chile_1275126_Sitios.pdf."
    )
    assert codes == []


def test_extract_candidate_file_names_from_question_preserves_pdf_names() -> None:
    file_names = extract_candidate_file_names_from_question(
        "Analiza todo el documento AI041.pdf y comparalo con AI041_Modificacion.pdf."
    )

    assert file_names == ["AI041.pdf", "AI041_Modificacion.pdf"]


def test_scope_resolver_uses_manual_scope_first() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Genera un analisis de los contratos seleccionados",
        user_id=7,
        file_ids=[104, 101],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "manual"
    assert resolution.file_ids == [104, 101]
    assert resolution.ignored_inferred_scope is False


def test_scope_resolver_requests_accessible_scope_queries() -> None:
    repository = _RecordingScopeRepository()
    resolver = QuestionScopeResolver(repository)
    resolution = resolver.resolve(
        question="Genera un analisis de RM797",
        user_id=7,
        file_ids=[104, 101, 102],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "inferred"
    assert resolution.file_ids == [101, 102]
    assert ("filter_file_ids_for_user", True) in repository.calls
    assert ("list_distinct_document_codes_for_user", True) in repository.calls


def test_scope_resolver_prefers_exact_pdf_filename_over_document_code() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Analiza todo el documento AI041.pdf y muestra sus campos clave.",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
    )

    assert resolution.scope_origin == "manual"
    assert resolution.scope_document_codes == []
    assert resolution.file_ids == [101]
    assert resolution.resolved_scope_file_count == 1
    assert resolution.ignored_inferred_scope is True


def test_scope_resolver_narrows_manual_scope_by_archive_slug() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Compara RM797_ID_1668 con RM797_ID_5515",
        user_id=7,
        file_ids=[104, 101, 102],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "metadata"
    assert resolution.scope_archive_slugs == ["RM797_ID_1668", "RM797_ID_5515"]
    assert resolution.file_ids == [101, 102]
    assert resolution.ignored_inferred_scope is False


def test_scope_resolver_narrows_manual_scope_by_document_code() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Genera un analisis de RM797",
        user_id=7,
        file_ids=[104, 101, 102],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "inferred"
    assert resolution.scope_document_codes == ["RM797"]
    assert resolution.file_ids == [101, 102]
    assert resolution.ignored_inferred_scope is False


def test_scope_resolver_infers_scope_from_multiple_codes() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Compara los contratos de RM797 con RM798",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "inferred"
    assert resolution.scope_document_codes == ["RM797", "RM798"]
    assert resolution.file_ids == [101, 102, 103, 104]


def test_extract_candidate_archive_slugs_from_question_dedupes_extensions() -> None:
    archive_slugs = extract_candidate_archive_slugs_from_question(
        "Compara RM797_ID_1668.zip con rm797_id_1668.pdf y RM797_ID_5515"
    )
    assert archive_slugs == ["RM797_ID_1668", "RM797_ID_5515"]


def test_parse_question_selectors_extracts_inline_metadata_file_and_column_scope() -> None:
    parsed = parse_question_selectors(
        question="@metadata /file:RM797_ID_1668 /col:Estado Contrato compara vigencia y cesion",
        available_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
        available_metadata_fields=["Estado Contrato", "Cesion a Terceros"],
    )

    assert parsed.cleaned_question == "compara vigencia y cesion"
    assert parsed.metadata_mode == "metadata_first"
    assert parsed.archive_slugs == ["RM797_ID_1668"]
    assert parsed.metadata_fields == ["Estado Contrato"]


def test_merge_question_selectors_combines_structured_and_inline_scope() -> None:
    merged = merge_question_selectors(
        question="/file:RM797_ID_1668 compara contratos",
        request_metadata_mode="auto",
        request_archive_slugs=["RM797_ID_5515"],
        request_metadata_fields=["Estado Contrato"],
        available_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
        available_metadata_fields=["Estado Contrato", "Forma de Pago"],
    )

    assert merged.cleaned_question == "compara contratos"
    assert merged.metadata_mode == "metadata_first"
    assert merged.archive_slugs == ["RM797_ID_5515", "RM797_ID_1668"]
    assert merged.metadata_fields == ["Estado Contrato"]


def test_build_effective_selector_question_allows_metadata_selector_only_lookup() -> None:
    merged = merge_question_selectors(
        question='/file:AI041_ID_49 /col:"Renta o Precio Vigente"',
        request_metadata_mode="auto",
        request_archive_slugs=[],
        request_metadata_fields=[],
        available_archive_slugs=["AI041_ID_49", "LA122_ID_3979"],
        available_metadata_fields=["Estado Contrato", "Renta o Precio Vigente"],
    )

    assert merged.cleaned_question == ""
    assert merged.metadata_mode == "metadata_first"
    assert merged.archive_slugs == ["AI041_ID_49"]
    assert merged.metadata_fields == ["Renta o Precio Vigente"]
    assert (
        build_effective_selector_question(merged)
        == "Muestra los valores de metadata seleccionados para los archivos seleccionados."
    )


def test_build_effective_selector_question_rejects_empty_non_selector_input() -> None:
    assert (
        build_effective_selector_question(
            ParsedQuestionSelectors(cleaned_question="", metadata_mode="auto", archive_slugs=[], metadata_fields=[])
        )
        == ""
    )


def test_scope_resolver_infers_scope_from_archive_slugs() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Compara RM797_ID_1668 con RM797_ID_5515",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
    )
    assert resolution.scope_origin == "metadata"
    assert resolution.scope_archive_slugs == ["RM797_ID_1668", "RM797_ID_5515"]
    assert resolution.file_ids == [101, 102]


def test_scope_resolver_applies_structured_archive_slug_scope_without_question_hints() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Profundiza en estos contratos",
        user_id=7,
        file_ids=[],
        archive_slugs=["rm797_id_5515.zip"],
        allow_inferred_scope=False,
    )

    assert resolution.scope_origin == "metadata"
    assert resolution.scope_archive_slugs == ["RM797_ID_5515"]
    assert resolution.file_ids == [102]


def test_scope_resolver_raises_when_structured_archive_slug_is_outside_manual_scope() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    with pytest.raises(ScopeResolutionError) as exc_info:
        resolver.resolve(
            question="Profundiza en este archivo",
            user_id=7,
            file_ids=[103, 104],
            archive_slugs=["RM797_ID_1668"],
            allow_inferred_scope=True,
        )

    assert exc_info.value.status_code == 404
    assert "RM797_ID_1668" in str(exc_info.value)


def test_scope_resolver_inherits_previous_conversation_scope_for_deictic_follow_up() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="De estos 5 sitios cuales son sus ultimos documentos firmados?",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
        conversation_file_ids=[101, 102, 101],
        conversation_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
    )
    assert resolution.scope_origin == "conversation"
    assert resolution.file_ids == [101, 102]
    assert resolution.scope_archive_slugs == ["RM797_ID_1668", "RM797_ID_5515"]


def test_scope_resolver_inherits_previous_conversation_scope_for_singular_follow_up() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Sobre ese mismo archivo, cita que dice la ultima modificacion sobre el acceso al terreno.",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
        conversation_file_ids=[101],
        conversation_archive_slugs=["RM797_ID_1668"],
    )
    assert resolution.scope_origin == "conversation"
    assert resolution.file_ids == [101]
    assert resolution.scope_archive_slugs == ["RM797_ID_1668"]


def test_scope_resolver_inherits_previous_scope_for_whole_document_follow_up() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Genera una lista valor de cada campo que consideres relevante para revisar todo el documento.",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
        conversation_file_ids=[102],
        conversation_archive_slugs=["RM797_ID_5515"],
    )

    assert resolution.scope_origin == "conversation"
    assert resolution.file_ids == [102]
    assert resolution.scope_archive_slugs == ["RM797_ID_5515"]


def test_scope_resolver_does_not_inherit_previous_scope_for_unrelated_question() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    resolution = resolver.resolve(
        question="Cuantos contratos estan vencidos?",
        user_id=7,
        file_ids=[],
        allow_inferred_scope=True,
        conversation_file_ids=[101, 102],
        conversation_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
    )
    assert resolution.scope_origin == "global"
    assert resolution.file_ids == []


def test_scope_resolver_raises_when_archive_slug_is_missing() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    with pytest.raises(ScopeResolutionError) as exc_info:
        resolver.resolve(
            question="Compara RM797_ID_1668 con RM999_ID_1",
            user_id=7,
            file_ids=[],
            allow_inferred_scope=True,
        )
    assert exc_info.value.status_code == 404
    assert "RM999_ID_1" in str(exc_info.value)


def test_scope_resolver_raises_when_any_code_is_missing() -> None:
    resolver = QuestionScopeResolver(_FakeScopeRepository())
    with pytest.raises(ScopeResolutionError) as exc_info:
        resolver.resolve(
            question="Compara RM797 con RM999",
            user_id=7,
            file_ids=[],
            allow_inferred_scope=True,
        )
    assert exc_info.value.status_code == 404
    assert "RM999" in str(exc_info.value)


def test_question_classifier_detects_temporal_question() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Â¿CuÃ¡nto tiempo le queda de vigencia al contrato? Si hoy es 20 de Marzo 2026",
    )
    assert result.question_class == "temporal"


def test_question_classifier_detects_document_inventory_request() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(question="Listame todos los documentos y archivos que tengo cargados.")
    assert result.question_class == "inventory"


def test_question_classifier_detects_associated_documents_request() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(question="Segun AI041_ID_49 cuales son sus documentos asociados?")
    assert result.question_class == "inventory"


def test_question_classifier_routes_key_value_document_follow_up_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Genera una lista valor de cada campo que consideres relevante para revisar todo el documento."
    )

    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_detects_analytics_question_without_accent_dependency() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(question="Cuantos contratos estan vencidos?")
    assert result.question_class == "analytics"


def test_question_classifier_detects_metadata_comparison() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Compara la metadata y las diferencias entre archivos RM797_ID_1668 y RM797_ID_5515"
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_global_metadata_analytics_to_analytics() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Usando toda la metadata cargada, que sitios tienen mas de un ID de contrato?"
    )
    assert result.question_class == "analytics"


def test_question_classifier_routes_region_metadata_count_to_analytics() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Segun la metadata cuantos sitios hay en la region metropolitana de Santiago?"
    )
    assert result.question_class == "analytics"


def test_question_classifier_routes_entel_aggregate_with_document_hint_to_analytics() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Usando toda la metadata cargada, cuantos contratos vigentes fueron firmados por ENTEL PCS? Si puedes, contrastalo con evidencia documental."
    )
    assert result.question_class == "analytics"


def test_question_classifier_does_not_route_metadata_document_differences_to_inventory() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Comparando metadata CSV contra los documentos procesados, en que contratos detectas diferencias "
            "relevantes en estado, cesion a terceros y acceso?"
        )
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_document_inventory_reasoning_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Que documentos integran el expediente y cuales son los que modifican el contrato base? "
            "Lista los nombres exactos de los PDF relevantes y cita la evidencia documental."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_routes_document_traceability_request_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "De donde fue extraido cada dato clave utilizado en la respuesta? "
            "Lista los nombres exactos de los PDF relevantes y cita la evidencia documental."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_routes_pdf_timeline_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Reconstruye la linea de tiempo documental de AI041_ID_49 usando AI041.pdf, "
            "AI041_Modificacin_1.pdf y AI041_Aclaracion_y_rectificacion.pdf."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_routes_mixed_metadata_and_documents_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Compara RM797_ID_1668 y RM797_ID_5515 usando metadata y documentos; "
            "indica tipo de documento, partes involucradas y fechas clave."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_routes_cross_archive_comparison_with_citations_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Compara RM797_ID_1668 y RM797_ID_5515 en estado, partes, forma de pago y fecha de termino. "
            "Usa metadata y documentos, y cita los PDFs mas relevantes."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_routes_versioned_clause_with_citation_to_versioned() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Para LA122_ID_3979, que dice el ultimo contrato vigente sobre cesion a terceros? "
            "Cita la clausula y el PDF."
        )
    )
    assert result.question_class == "versioned"


def test_question_classifier_routes_follow_up_latest_signed_documents_to_versioned() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="De estos 5 sitios cuales son sus ultimos documentos firmados?"
    )
    assert result.question_class == "versioned"


def test_question_classifier_routes_metadata_validation_with_documents_to_metadata_comparison() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Usando metadata y documentos de RM797_ID_1668, valida si el Estado Contrato es desconocido "
            "y confirma la Forma de Pago con evidencia."
        )
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_ocr_content_request_to_document_synthesis() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Usar @metadata y /file:AI041_ID_49. Segun el OCR del documento, resume de que trata "
            "el contrato y menciona las partes principales, la direccion o sitio y la renta si aparece."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_question_classifier_does_not_confuse_confirma_with_firma() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Usa la metadata para encontrar el file RM797_ID_1668 y confirma si el Estado Contrato "
            "es vigente; además indica la Comuna y la Dirección."
        )
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_beneficiary_rut_question_to_metadata_comparison() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Para RM797_ID_5515, quien recibe la renta y cual es su RUT?"
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_dynamic_archive_metadata_lookup_to_metadata_comparison() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Para RM797_ID_5515, cual es el Segmento Comercial y el Responsable Comercial?"
    )
    assert result.question_class == "metadata_comparison"


def test_question_classifier_routes_dynamic_metadata_aggregate_to_analytics() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question="Segun la metadata, cuantos segmentos comerciales hay?"
    )
    assert result.question_class == "analytics"


def test_question_requires_visual_grounding_uses_whole_tokens() -> None:
    assert question_requires_visual_grounding("Muéstrame la firma del representante.") is True
    assert question_requires_visual_grounding("Confirma la vigencia del contrato y la dirección.") is False


def test_question_requires_visual_grounding_does_not_trigger_for_pdf_mentions() -> None:
    assert question_requires_visual_grounding("Compara AI041.pdf y AI041_Modificacin_1.pdf.") is False


def test_question_classifier_does_not_route_signature_dates_to_visual_consistency() -> None:
    classifier = QuestionClassifier()
    result = classifier.classify(
        question=(
            "Compara AI041.pdf, AI041_Modificacin_1.pdf y AI041_Aclaracion_y_rectificacion.pdf dentro de "
            "AI041_ID_49; resume cambios de fechas de firma, notaria, repertorio y representantes, y cita cada PDF."
        )
    )
    assert result.question_class == "exhaustive_synthesis"


def test_extract_latest_conversation_scope_from_messages_uses_latest_usable_scope_and_derives_archive_slugs() -> None:
    extracted = _extract_latest_conversation_scope_from_messages(
        conversation_messages=[
            {
                "session_id": 51,
                "turn_index": 1,
                "role": "assistant",
                "retrieval_metadata": {
                    "scope_file_ids": [101],
                    "question_class": "metadata_comparison",
                    "answer_mode": "facts-first",
                },
            },
            {
                "session_id": 88,
                "turn_index": 2,
                "role": "assistant",
                "retrieval_metadata": {
                    "scope_file_ids": [101, 102],
                    "question_class": "analytics",
                    "answer_mode": "facts-first",
                },
            },
            {
                "session_id": 88,
                "turn_index": 2,
                "role": "user",
                "content": "De estos 5 sitios cuales son sus ultimos documentos firmados?",
            },
        ],
        archive_slug_map_resolver=lambda file_ids: {
            101: "RM797_ID_1668",
            102: "RM797_ID_5515",
        },
    )
    assert extracted["conversation_scope_file_ids"] == [101, 102]
    assert extracted["conversation_scope_archive_slugs"] == ["RM797_ID_1668", "RM797_ID_5515"]
    assert extracted["conversation_scope_turn_index"] == 2
    assert extracted["conversation_scope_question_class"] == "analytics"
    assert extracted["conversation_scope_answer_mode"] == "facts-first"


def test_retrieval_pipeline_treats_conversation_scope_as_scoped_origin() -> None:
    assert RetrievalPipelineService._is_scoped_origin("conversation") is True


def test_extract_explicit_file_references_preserves_each_pdf_once() -> None:
    assert extract_explicit_file_references(
        "Compara AI041.pdf, AI041_Modificacin_1.pdf y AI041.pdf dentro del mismo folio."
    ) == [
        "AI041.pdf",
        "AI041_Modificacin_1.pdf",
    ]


def test_build_file_group_key_prefers_primary_identifier() -> None:
    value = build_file_group_key(
        primary_identifier="RM797-5515",
        secondary_identifier="RM797",
        primary_subject="Entel",
        secondary_subject="Transam",
    )
    assert value == "primary:RM797-5515"


def test_build_file_group_key_uses_secondary_identifier_and_subjects() -> None:
    value = build_file_group_key(
        primary_identifier=None,
        secondary_identifier="RM797",
        primary_subject="Entel",
        secondary_subject="Transam",
    )
    assert value == "secondary:RM797|primary_subject:ENTEL|secondary_subject:TRANSAM"


def test_build_file_group_key_truncates_on_utf8_byte_boundary() -> None:
    value = build_file_group_key(
        primary_identifier=None,
        secondary_identifier="ESTBI044",
        primary_subject="Ñ" * 200,
        secondary_subject="Á" * 200,
    )
    assert value is not None
    assert len(value.encode("utf-8")) <= 256


def test_extract_secondary_identifier_matches_rm_token_without_capture_group() -> None:
    assert _extract_secondary_identifier("Contrato sitio RM797 firmado por las partes.") == "RM797"


def test_parse_reference_date_from_question() -> None:
    resolved = QuestionFactResolver._parse_reference_date(
        "Â¿CuÃ¡nto tiempo le queda de vigencia al contrato? Si hoy es 20 de Marzo 2026"
    )
    assert resolved == date(2026, 3, 20)


def test_question_fact_resolver_answers_metadata_lookup_from_csv() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Id": 5515,
                                "Forma de Pago": "Deposito",
                                "Estado Contrato": "Terminado",
                                "Comuna": "Las Condes",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Filtra por file RM797_ID_5515 y revisa la Forma de Pago reportada en metadata.",
        user_id=7,
        file_ids=[102],
    )

    assert resolution.narrowed_file_ids == [102]
    assert resolution.answer_override == (
        "En la metadata de RM797_ID_5515:\n\n"
        "| Campo | Valor |\n"
        "| --- | --- |\n"
        "| Forma de Pago | Deposito |"
    )
    assert resolution.facts_used_count == 1
    assert "RM797_ID_5515: Forma de Pago=Deposito" in resolution.fact_context_text


def test_question_fact_resolver_keeps_metadata_first_file_inventory_open_for_document_modification_reasoning() -> None:
    class _MetadataAndInventoryRepository(_FakeArchiveMetadataFileRepository):
        def get_archive_slug_map_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> dict[int, str]:
            del user_id, include_shared
            allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
            return {
                int(row.get("file_id") or 0): str(row.get("archive_slug") or "")
                for row in self.file_rows
                if int(row.get("file_id") or 0) in allowed and str(row.get("archive_slug") or "").strip()
            }

    file_rows = [
        {
            "file_id": file_id,
            "archive_slug": "LA122_ID_3979",
            "file_input_file_name": file_name,
            "file_code": "LA122",
            "file_state": 3,
            "file_page_count": pages,
        }
        for file_id, file_name, pages in (
            (201, "LA122_Contrato_Base.pdf", 10),
            (202, "LA122_Modificacion_1.pdf", 12),
            (203, "LA122_Modificacion_2.pdf", 8),
            (204, "LA122_Anexo_Canon.pdf", 4),
            (205, "LA122_Acta_Entrega.pdf", 3),
            (206, "LA122_Carta_Aviso.pdf", 2),
        )
    ]
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_MetadataAndInventoryRepository(
            [
                {
                    "file_id": 201,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Codigo de Sitio": "LA122",
                                "Tipo de Contrato": "Arriendo",
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ],
            file_rows=file_rows,
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Que documentos integran el expediente y cuales modifican el contrato base?",
        user_id=7,
        file_ids=[201, 202, 203, 204, 205, 206],
        metadata_mode="metadata_first",
        archive_slugs=["LA122_ID_3979"],
    )

    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.narrowed_file_ids == [201, 202, 203, 204, 205, 206]
    assert "Document inventory context" in resolution.fact_context_text
    assert "LA122_Contrato_Base.pdf" in resolution.fact_context_text
    assert "LA122_Modificacion_1.pdf" in resolution.fact_context_text
    assert "LA122_Carta_Aviso.pdf" in resolution.fact_context_text


def test_question_fact_resolver_keeps_document_grounded_scope_within_requested_file_ids() -> None:
    class _FakeFactsRepository:
        def list_group_ids_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> list[int]:
            assert user_id == 7
            assert file_ids == [101, 102]
            assert include_shared is True
            return [11, 12]

        def list_file_ids_for_group_ids(
            self,
            *,
            user_id: int,
            group_ids: list[int],
            current_only: bool,
            include_shared: bool = False,
        ) -> list[int]:
            assert user_id == 7
            assert group_ids == [11, 12]
            assert current_only is False
            assert include_shared is True
            return [101, 999, 102, 998]

    resolver = QuestionFactResolver(
        repository=_FakeFactsRepository(),
        file_repository=_FakeArchiveMetadataFileRepository([]),
    )

    resolution = resolver.resolve(
        question_class="exhaustive_synthesis",
        question="Compara RM797_ID_1668 y RM797_ID_5515 usando metadata y documentos.",
        user_id=7,
        file_ids=[101, 102],
    )

    assert resolution.narrowed_file_ids == [101, 102]
    assert resolution.file_group_ids == [11, 12]
    assert resolution.facts_used_count == 2


def test_question_fact_resolver_returns_document_inventory_from_files_repository() -> None:
    file_repository = _RecordingInventoryFileRepository(
        [
            {
                "file_id": 201,
                "archive_slug": "AI041_ID_49",
                "file_input_file_name": "AI041.pdf",
                "file_code": "AI041",
                "file_state": 3,
                "file_page_count": 18,
            },
            {
                "file_id": 202,
                "archive_slug": "RM797_ID_1668",
                "file_input_file_name": "RM797_contrato.pdf",
                "file_code": "RM797",
                "file_state": 2,
                "file_page_count": 12,
            },
        ]
    )
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=file_repository,
    )

    resolution = resolver.resolve(
        question_class="inventory",
        question="Listame todos los documentos que tengo disponibles.",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "Documentos disponibles (2)" in resolution.answer_override
    assert "| # | Archivo | Documento | Codigo | Estado | Paginas |" in resolution.answer_override
    assert "| 1 | AI041_ID_49 | AI041.pdf | AI041 | completed | 18 |" in resolution.answer_override
    assert "| 2 | RM797_ID_1668 | RM797_contrato.pdf | RM797 | processing | 12 |" in resolution.answer_override
    assert resolution.facts_used_count == 2
    assert file_repository.include_shared_calls == [True]


def test_question_fact_resolver_prefers_associated_documents_over_inherited_metadata_field() -> None:
    file_repository = _FakeArchiveMetadataFileRepository(
        [
            {
                "file_id": 201,
                "archive_slug": "AI041_ID_49",
                "metadata_json": json.dumps(
                    {
                        "file": "AI041_ID_49",
                        "fields": {
                            "Renta o Precio Vigente": "420",
                        },
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        file_rows=[
            {
                "file_id": 201,
                "archive_slug": "AI041_ID_49",
                "file_input_file_name": "AI041.pdf",
                "file_code": "AI041",
                "file_state": 3,
                "file_page_count": 18,
            },
            {
                "file_id": 202,
                "archive_slug": "AI041_ID_49",
                "file_input_file_name": "AI041_Modificacion_1.pdf",
                "file_code": "AI041",
                "file_state": 3,
                "file_page_count": 6,
            },
        ],
    )
    resolver = QuestionFactResolver(repository=object(), file_repository=file_repository)

    resolution = resolver.resolve(
        question_class="extractive",
        question="Segun AI041_ID_49 cuales son sus documentos asociados?",
        user_id=7,
        file_ids=[201, 202],
        metadata_mode="auto",
        archive_slugs=["AI041_ID_49"],
        metadata_fields=["Renta o Precio Vigente"],
    )

    assert resolution.answer_override is not None
    assert "Documentos asociados a AI041_ID_49 (2)" in resolution.answer_override
    assert "AI041.pdf" in resolution.answer_override
    assert "AI041_Modificacion_1.pdf" in resolution.answer_override
    assert "Renta o Precio Vigente" not in resolution.answer_override
    assert "420" not in resolution.answer_override


def test_question_fact_resolver_keeps_inventory_as_context_when_question_requires_document_reasoning() -> None:
    file_repository = _RecordingInventoryFileRepository(
        [
            {
                "file_id": 201,
                "archive_slug": "LA122_ID_3979",
                "file_input_file_name": "LA122.PDF",
                "file_code": "LA122",
                "file_state": 3,
                "file_page_count": 10,
            },
            {
                "file_id": 202,
                "archive_slug": "LA122_ID_3979",
                "file_input_file_name": "LA122_Modificacion.pdf",
                "file_code": "LA122",
                "file_state": 3,
                "file_page_count": 12,
            },
            {
                "file_id": 203,
                "archive_slug": "LA122_ID_3979",
                "file_input_file_name": "LA122_Modificacion_2.pdf",
                "file_code": "LA122",
                "file_state": 3,
                "file_page_count": 8,
            },
        ]
    )
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=file_repository,
    )

    resolution = resolver.resolve(
        question_class="inventory",
        question=(
            "Que documentos integran el expediente y cuales son los que modifican el contrato base? "
            "Lista los nombres exactos de los PDF relevantes y cita la evidencia documental."
        ),
        user_id=7,
        file_ids=[201, 202, 203],
    )

    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.narrowed_file_ids == [201, 202, 203]
    assert "Documentos disponibles (3)" in resolution.fact_context_text
    assert "LA122_Modificacion.pdf" in resolution.fact_context_text


def test_graph_synthesis_per_document_uses_inventory_for_missing_evidence() -> None:
    synthesis = GraphSynthesis(provider=object())
    evidence = [
        _make_evidence_item(
            file_id=201,
            file_name="LA122.PDF",
            page_id=1,
            source_number=1,
            summary_text="Contrato base con condiciones originales.",
        ),
        _make_evidence_item(
            file_id=203,
            file_name="LA122_Modificacion_2.pdf",
            page_id=3,
            source_number=3,
            summary_text="Segunda modificacion con cambios de renta.",
        ),
    ]
    fact_context = "\n".join(
        [
            "Document inventory context:",
            "- file_id=201 archive=LA122_ID_3979 file=LA122.PDF status=completed pages=10",
            "- file_id=202 archive=LA122_ID_3979 file=LA122_Modificacion.pdf status=completed pages=12",
            "- file_id=203 archive=LA122_ID_3979 file=LA122_Modificacion_2.pdf status=completed pages=8",
        ]
    )

    result = synthesis.synthesize(
        question="Que documentos integran el expediente?",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=3,
        fact_context=fact_context,
        question_class="exhaustive_synthesis",
    )

    assert "Inventario documental completo" in result.answer_text
    assert "LA122.PDF" in result.answer_text
    assert "LA122_Modificacion.pdf" in result.answer_text
    assert "LA122_Modificacion_2.pdf" in result.answer_text
    assert "sin evidencia OCR suficiente" in result.answer_text
    assert "Se resumieron 2 de 3" not in result.answer_text


def test_graph_synthesis_per_document_formats_document_inventory_as_readable_markdown() -> None:
    synthesis = GraphSynthesis(provider=object())
    evidence = [
        _make_evidence_item(
            file_id=201,
            file_name="LA122.PDF",
            page_id=1,
            source_number=1,
            summary_text=(
                "Contrato base con condiciones originales sobre renta, plazo y autorizaciones. "
                "DECIMO CUARTO: texto OCR largo que no debe convertir la respuesta en una sabana."
            ),
        ),
        _make_evidence_item(
            file_id=202,
            file_name="LA122_Modificacion.pdf",
            page_id=2,
            source_number=2,
            summary_text="Modificacion del contrato de arrendamiento que ajusta condiciones del contrato base.",
        ),
    ]
    fact_context = "\n".join(
        [
            "Archive metadata context:",
            (
                "LA122_ID_3979: Estado Contrato=Vigente; Estado Actividad=Activo; "
                "Revision Final=REVISADO OK; Renta o Precio Vigente=504; Tipo de Moneda=UF; "
                "Fecha de Inicio de Vigencia del Contrato=01/08/2025; "
                "Fecha de Termino del Contrato=01/08/2027; Nombre de Propietario Principal=TRANSPORTES COSTANERA S.A."
            ),
            "Document inventory context:",
            "- file_id=201 archive=LA122_ID_3979 file=LA122.PDF status=completed pages=10",
            "- file_id=202 archive=LA122_ID_3979 file=LA122_Modificacion.pdf status=completed pages=12",
        ]
    )

    result = synthesis.synthesize(
        question="Que documentos integran el expediente y cuales modifican el contrato base?",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=2,
        fact_context=fact_context,
        question_class="exhaustive_synthesis",
    )

    assert "## Resumen" in result.answer_text
    assert "## Metadata clave" in result.answer_text
    assert "## Documentos del expediente" in result.answer_text
    assert "## Documentos que modifican o complementan el contrato base" in result.answer_text
    assert "**LA122.PDF**" in result.answer_text
    assert "**LA122_Modificacion.pdf**" in result.answer_text
    assert "[1]" not in result.answer_text
    assert "[2]" not in result.answer_text
    assert "sabana" not in result.answer_text
    assert result.citation_source_numbers == [1, 2]


def test_graph_synthesis_per_document_surfaces_broad_metadata_context() -> None:
    synthesis = GraphSynthesis(provider=object())
    evidence = [
        _make_evidence_item(
            file_id=201,
            file_name="LA122.PDF",
            page_id=1,
            source_number=1,
            summary_text="Contrato base con condiciones originales.",
        )
    ]
    fields_before_key_facts = "; ".join(f"Campo Auxiliar {index}=valor {index}" for index in range(1, 16))
    fact_context = "\n".join(
        [
            "Archive metadata context:",
            (
                "LA122_ID_3979: "
                f"{fields_before_key_facts}; "
                "Renta o Precio Vigente=504; Tipo de Moneda=UF; Periodo de Pago=Anual; "
                "Fecha de Inicio de Vigencia del Contrato=01/08/2025; "
                "Fecha de Termino del Contrato=01/08/2027"
            ),
            "Document inventory context:",
            "- file_id=201 archive=LA122_ID_3979 file=LA122.PDF status=completed pages=10",
        ]
    )

    result = synthesis.synthesize(
        question="Cual es el instrumento vigente que gobierna cada variable critica del expediente?",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=1,
        fact_context=fact_context,
        question_class="exhaustive_synthesis",
    )

    assert "Metadata estructurada" in result.answer_text
    assert "504" in result.answer_text
    assert "UF" in result.answer_text
    assert "01/08/2025" in result.answer_text
    assert "01/08/2027" in result.answer_text


def test_graph_synthesis_per_document_surfaces_resolved_metadata_facts() -> None:
    synthesis = GraphSynthesis(provider=object())
    evidence = [
        _make_evidence_item(
            file_id=301,
            file_name="RM797_-_Contrato_2.pdf",
            page_id=1,
            source_number=1,
            summary_text="Contrato con evidencia documental.",
        )
    ]

    result = synthesis.synthesize(
        question="De donde fue extraido cada dato clave utilizado en la respuesta?",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=1,
        fact_context=(
            "Resolved metadata facts:\n"
            "RM797_ID_5515: Estado Contrato=Terminado; Estado Actividad=Inactivo"
        ),
        question_class="exhaustive_synthesis",
    )

    assert "Metadata estructurada" in result.answer_text
    assert "Estado Contrato=Terminado" in result.answer_text
    assert "Estado Actividad=Inactivo" in result.answer_text


def test_graph_synthesis_per_document_keeps_representative_window() -> None:
    synthesis = GraphSynthesis(provider=object())
    filler = " ".join(f"texto{index}" for index in range(180))
    representative_text = (
        f"{filler} comparecen: por una parte SOCIEDAD TRANSPORTES COSTANERA S.A., "
        "representada por don MARIO CARLOS PACHECO VAZQUEZ y por dona "
        "JANETTE LUCILA MANSILLA TOLEDO; y por la otra ENTEL PCS TELECOMUNICACIONES S.A., "
        "representada por don FRANCISCO JAVIER SPRENGER ARROYO. "
        f"{filler}"
    )
    evidence = [
        _make_evidence_item(
            file_id=202,
            file_name="LA122_Modificacion.pdf",
            page_id=2,
            source_number=2,
            summary_text=representative_text,
        )
    ]

    result = synthesis.synthesize(
        question="Que personas o representantes aparecen con facultades para firmar?",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=1,
        fact_context="",
        question_class="exhaustive_synthesis",
    )

    assert "MARIO CARLOS PACHECO VAZQUEZ" in result.answer_text
    assert "JANETTE LUCILA MANSILLA TOLEDO" in result.answer_text
    assert "FRANCISCO JAVIER SPRENGER ARROYO" in result.answer_text


def test_graph_synthesis_per_document_metadata_question_uses_llm_not_inventory_answer() -> None:
    class _ResolvedConfig:
        model_id = "fake-model"

    class _FakeProvider:
        def __init__(self) -> None:
            self.prompts: list[str] = []

        def is_available(self) -> bool:
            return True

        def invoke_text(self, *, prompt: str, model_id: str | None = None) -> str:
            del model_id
            self.prompts.append(prompt)
            return (
                "ANSWER:\n"
                "Metadata resuelta:\n\n"
                "| Archivo | Estado Contrato | Renta o Precio Vigente |\n"
                "| --- | --- | --- |\n"
                "| RM797_ID_1668 | Vigente | 442 |\n"
                "| RM797_ID_5515 | Terminado | 45 |\n\n"
                "Con esa metadata como contexto, la evidencia documental permite responder "
                "la pregunta sin mostrar el inventario interno del expediente.\n"
                "EXECUTIVE_SUMMARY: Respuesta mixta generada con metadata y evidencia documental.\n"
                "KEY_POINTS:\n"
                "- Metadata usada como contexto.\n"
                "- Evidencia documental usada para la conclusion.\n"
                "OBLIGATIONS:\n"
                "CITATIONS: 1,2"
            )

        def resolve_config(self) -> _ResolvedConfig:
            return _ResolvedConfig()

    provider = _FakeProvider()
    synthesis = GraphSynthesis(provider=provider)

    result = synthesis.synthesize(
        question="¿Hay penalización por pago atrasado de renta? RM797",
        evidence=[
            _make_evidence_item(
                file_id=301,
                file_name="RM797-Contrato_2.pdf",
                page_id=8001,
                page_number=8,
                source_number=1,
                summary_text="Clausula con condiciones de pago y efectos del atraso.",
            ),
            _make_evidence_item(
                file_id=302,
                file_name="RM797_Rectificacion.pdf",
                page_id=3001,
                page_number=3,
                source_number=2,
                summary_text="Rectificacion relacionada con antecedentes del mismo expediente.",
            ),
        ],
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=2,
        fact_context=(
            "Resolved metadata facts:\n"
            "RM797_ID_1668: Estado Contrato=Vigente; Renta o Precio Vigente=442\n"
            "RM797_ID_5515: Estado Contrato=Terminado; Renta o Precio Vigente=45\n"
            "Document inventory context:\n"
            "- file_id=301 archive=RM797_ID_1668 file=RM797-Contrato_2.pdf status=completed pages=12\n"
            "- file_id=302 archive=RM797_ID_5515 file=RM797_Rectificacion.pdf status=completed pages=3"
        ),
        question_class="metadata_comparison",
    )

    assert provider.prompts
    assert result.model_used == "langgraph-oci-synthesis:fake-model"
    assert "Metadata resuelta" in result.answer_text
    assert "Inventario documental completo" not in result.answer_text
    assert "## Documentos del expediente" not in result.answer_text
    assert "Lectura OCR" not in result.answer_text
    assert result.citation_source_numbers == [1, 2]


def test_graph_synthesis_retries_tabular_request_when_first_answer_lacks_markdown_table() -> None:
    class _ResolvedConfig:
        model_id = "fake-model"

    class _FakeProvider:
        def __init__(self) -> None:
            self.prompts: list[str] = []
            self.responses = [
                (
                    "A continuacion se presenta una tabla con campos relevantes, pero sin tabla real.\n"
                    "EXECUTIVE_SUMMARY: resumen interno\n"
                    "KEY_POINTS:\n"
                    "- punto\n"
                    "CITATIONS: 1,2"
                ),
                (
                    "ANSWER:\n"
                    "| Campo | Valor | Fuente | Nota |\n"
                    "| --- | --- | --- | --- |\n"
                    "| Tipo de Documento | Contrato de arrendamiento y servidumbres | AI041.pdf - page 1 | Extraido del encabezado contractual |\n"
                    "| Renta Anual | 200 UF | AI041.pdf - page 7 | Pago anual anticipado |\n"
                    "EXECUTIVE_SUMMARY: Tabla de campos clave generada desde evidencia OCR.\n"
                    "KEY_POINTS:\n"
                    "- Tipo de documento identificado\n"
                    "- Renta anual identificada\n"
                    "OBLIGATIONS:\n"
                    "- Pago anual anticipado\n"
                    "CITATIONS: 1,2"
                ),
            ]

        def is_available(self) -> bool:
            return True

        def invoke_text(self, *, prompt: str, model_id: str | None = None) -> str:
            del model_id
            self.prompts.append(prompt)
            return self.responses.pop(0)

        def resolve_config(self) -> _ResolvedConfig:
            return _ResolvedConfig()

    provider = _FakeProvider()
    synthesis = GraphSynthesis(provider=provider)

    result = synthesis.synthesize(
        question=(
            "Analiza todo el documento AI041.pdf y muestra todos los campos que consideres "
            "relevantes en una tabla con sus referencias por pagina."
        ),
        evidence=[
            _make_evidence_item(
                file_id=701,
                file_name="AI041.pdf",
                page_id=9001,
                page_number=1,
                source_number=1,
                summary_text="Contrato de arrendamiento y servidumbres.",
            ),
            _make_evidence_item(
                file_id=701,
                file_name="AI041.pdf",
                page_id=9007,
                page_number=7,
                source_number=2,
                summary_text="Renta anual de 200 Unidades de Fomento, pagadera en forma anual anticipada.",
            ),
        ],
        strategy="deep-reasoning",
        question_class="exhaustive_synthesis",
    )

    assert len(provider.prompts) == 2
    assert "faltaba una tabla Markdown valida" in provider.prompts[1]
    assert "| Campo | Valor | Fuente | Nota |" in result.answer_text
    assert "| Renta Anual | 200 UF | AI041.pdf - page 7 | Pago anual anticipado |" in result.answer_text
    assert "EXECUTIVE_SUMMARY" not in result.answer_text
    assert result.executive_summary == "Tabla de campos clave generada desde evidencia OCR."
    assert result.citation_source_numbers == [1, 2]


def test_graph_synthesis_extracts_raw_answer_without_internal_sections() -> None:
    raw_text = (
        "A continuacion se presenta el resultado solicitado.\n"
        "EXECUTIVE_SUMMARY: este texto no debe mostrarse dentro de ANSWER\n"
        "CITATIONS: 1"
    )

    assert (
        GraphSynthesis._extract_answer_section_or_raw(raw_text)
        == "A continuacion se presenta el resultado solicitado."
    )


def test_repair_document_file_name_handles_resciliacion_loss() -> None:
    repaired = repair_document_file_name("RM797_-_Resciliacin_Arrendamiento_Finiquito_y_Pago.pdf")

    assert repaired == "RM797_-_Resciliacion_Arrendamiento_Finiquito_y_Pago.pdf"


def test_markdown_selector_expected_terms_normalize_repaired_filenames_and_slashes() -> None:
    from apps.backend.tests.run_rag_markdown_selector_battery import _expected_terms_report

    matched, missing = _expected_terms_report(
        answer_text=(
            "Estado Contrato: Terminado; Estado Actividad: Inactivo. "
            "Documento: AI041_Carta_Aviso_Cesion_Contrato_Alba_ATC.pdf."
        ),
        expected_terms=(
            "Terminado/Inactivo",
            "AI041_Carta_Aviso_Cesin_Contrato_Alba_ATC.pdf",
        ),
    )

    assert matched == [
        "Terminado/Inactivo",
        "AI041_Carta_Aviso_Cesin_Contrato_Alba_ATC.pdf",
    ]
    assert missing == []


def test_hybrid_answer_tool_keeps_all_evidence_for_per_document_mode() -> None:
    class _FakeSynthesisAgent:
        def __init__(self) -> None:
            self.calls: list[dict[str, object]] = []

        def run(self, **kwargs) -> LLMResult:
            self.calls.append(dict(kwargs))
            evidence = list(kwargs["evidence"])
            names = ", ".join(item.file_name for item in evidence)
            return LLMResult(
                answer_text=names,
                executive_summary=names,
                key_points=[names],
                obligations=[],
                citation_source_numbers=[int(item.source_number) for item in evidence],
                model_used="fake-synthesis",
            )

    class _FakePageVisionTool:
        def analyze(self, **kwargs) -> VisualInspectionResult:
            raise AssertionError("visual analysis is not expected in this test")

    synthesis_agent = _FakeSynthesisAgent()
    tool = HybridAnswerTool(
        settings=Settings(_env_file=None, ANSWER_MAX_EVIDENCE=3),
        page_vision_tool=_FakePageVisionTool(),
        synthesis_agent=synthesis_agent,
    )
    evidence = [
        _make_evidence_item(file_id=file_id, file_name=f"doc-{file_id}.pdf", source_number=file_id)
        for file_id in range(1, 7)
    ]

    result = tool.answer(
        question="Lista los documentos seleccionados.",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="per_document",
        selected_docs_count=6,
        question_class="exhaustive_synthesis",
    )

    assert len(synthesis_agent.calls) == 1
    assert len(synthesis_agent.calls[0]["evidence"]) == 6
    assert "doc-6.pdf" in result.llm_result.answer_text
    assert result.llm_result.citation_source_numbers == [1, 2, 3, 4, 5, 6]


def test_hybrid_answer_tool_keeps_all_evidence_for_explicit_full_document_request() -> None:
    class _FakeSynthesisAgent:
        def __init__(self) -> None:
            self.calls: list[dict[str, object]] = []

        def run(self, **kwargs) -> LLMResult:
            self.calls.append(dict(kwargs))
            evidence = list(kwargs["evidence"])
            pages = ", ".join(str(item.page_number) for item in evidence)
            return LLMResult(
                answer_text=pages,
                executive_summary=pages,
                key_points=[pages],
                obligations=[],
                citation_source_numbers=[int(item.source_number) for item in evidence],
                model_used="fake-synthesis",
            )

    class _FakePageVisionTool:
        def analyze(self, **kwargs) -> VisualInspectionResult:
            raise AssertionError("visual analysis is not expected in this test")

    synthesis_agent = _FakeSynthesisAgent()
    tool = HybridAnswerTool(
        settings=Settings(_env_file=None, ANSWER_MAX_EVIDENCE=3),
        page_vision_tool=_FakePageVisionTool(),
        synthesis_agent=synthesis_agent,
    )
    evidence = [
        _make_evidence_item(
            file_id=701,
            file_name="AI041.pdf",
            page_id=9000 + page_number,
            source_number=page_number,
            page_number=page_number,
            summary_text=f"Texto OCR pagina {page_number}.",
        )
        for page_number in range(1, 7)
    ]

    result = tool.answer(
        question="Analiza todo el documento AI041.pdf y muestrame una lista completa clave valor.",
        evidence=evidence,
        strategy="deep-reasoning",
        summary_mode="default",
        selected_docs_count=1,
        question_class="exhaustive_synthesis",
    )

    assert len(synthesis_agent.calls) == 1
    assert len(synthesis_agent.calls[0]["evidence"]) == 6
    assert result.llm_result.answer_text == "1, 2, 3, 4, 5, 6"
    assert any("Cobertura de documento completo: 6 paginas" in note for note in result.confidence_notes)


def test_hybrid_answer_tool_prepends_metadata_table_for_mixed_document_answer() -> None:
    class _FakeSynthesisAgent:
        def run(self, **kwargs) -> LLMResult:
            del kwargs
            return LLMResult(
                answer_text=(
                    "## Metadata clave\n"
                    "Metadata estructurada priorizada desde el CSV.\n\n"
                    "No hay evidencia suficiente en los documentos provistos que especifique una penalizacion "
                    "por pago atrasado de renta."
                ),
                executive_summary="No se encontro penalizacion documental.",
                key_points=["No se encontro penalizacion documental."],
                obligations=[],
                citation_source_numbers=[1, 2],
                model_used="fake-synthesis",
            )

    class _FakePageVisionTool:
        def analyze(self, **kwargs) -> VisualInspectionResult:
            raise AssertionError("visual analysis is not expected in this test")

    tool = HybridAnswerTool(
        settings=Settings(_env_file=None, ANSWER_MAX_EVIDENCE=5),
        page_vision_tool=_FakePageVisionTool(),
        synthesis_agent=_FakeSynthesisAgent(),
    )

    result = tool.answer(
        question="Hay penalizacion por pago atrasado de renta? RM797",
        evidence=[
            _make_evidence_item(
                file_id=101,
                file_name="RM797-Contrato_2.pdf",
                source_number=1,
                page_number=8,
                summary_text=(
                    "El contrato detalla el valor de la renta y la forma de pago, pero este extracto OCR "
                    "no describe multas por retraso."
                ),
            ),
            _make_evidence_item(
                file_id=102,
                file_name="RM797_Rectificacion.pdf",
                source_number=2,
                page_number=3,
                summary_text=(
                    "La rectificacion confirma antecedentes del contrato y mantiene referencias de pago, "
                    "sin describir una penalizacion por mora."
                ),
            ),
        ],
        strategy="fast-grounded",
        question_class="metadata_comparison",
        fact_context_text=(
            "Resolved metadata facts:\n"
            "RM797_ID_1668: Renta o Precio Vigente=442; Pago Anticipado=NO; Periodo de Pago=Anual\n"
            "RM797_ID_5515: Renta o Precio Vigente=45; Pago Anticipado=NO; Periodo de Pago=Mensual\n"
            "Archive metadata context:\n"
            "RM797_ID_1668: Estado Contrato=Vigente"
        ),
    )

    assert result.llm_result.answer_text.startswith(
        "Metadata resuelta:\n\n"
        "| Archivo | Renta o Precio Vigente | Pago Anticipado | Periodo de Pago |\n"
        "| --- | --- | --- | --- |\n"
        "| RM797_ID_1668 | 442 | NO | Anual |\n"
        "| RM797_ID_5515 | 45 | NO | Mensual |"
    )
    assert "No hay evidencia suficiente en los documentos provistos" in result.llm_result.answer_text
    assert result.llm_result.citation_source_numbers == [1, 2]
    assert any("Metadata table" in note for note in result.confidence_notes)


def test_question_fact_resolver_answers_global_multi_contract_site_question_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "LA122_ID_18467",
                    "metadata_json": json.dumps(
                        {"file": "LA122_ID_18467", "fields": {"Codigo de Sitio": "LA122", "Id": "18467"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {"file": "LA122_ID_3979", "fields": {"Codigo de Sitio": "LA122", "Id": "3979"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 103,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {"file": "RM797_ID_1668", "fields": {"Codigo de Sitio": "RM797", "Id": "1668"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 104,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {"file": "RM797_ID_5515", "fields": {"Codigo de Sitio": "RM797", "Id": "5515"}},
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Usando toda la metadata cargada, que sitios tienen mas de un ID de contrato?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "LA122: 18467, 3979" in resolution.answer_override
    assert "RM797: 1668, 5515" in resolution.answer_override
    assert resolution.facts_used_count == 4


def test_question_fact_resolver_prefers_requested_metadata_fields_over_generic_aggregate_counts() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "ESTAN041_ID_14010",
                    "metadata_json": json.dumps(
                        {
                            "file": "ESTAN041_ID_14010",
                            "fields": {
                                "Nombre de Propietario Principal": "MINISTERIO DE BIENES NACIONALES",
                                "Nombre Beneficiario": "MINISTERIO DE BIENES NACIONALES",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "ZU163_ID_3630",
                    "metadata_json": json.dumps(
                        {
                            "file": "ZU163_ID_3630",
                            "fields": {
                                "Nombre de Propietario Principal": "ATC SITIOS CHILE S.A.",
                                "Nombre Beneficiario": "MINISTERIO DE BIENES NACIONALES",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Que contratos comparten el mismo propietario principal o beneficiario?",
        user_id=7,
        file_ids=[101, 102],
        metadata_fields=["Nombre de Propietario Principal", "Nombre Beneficiario"],
    )

    assert resolution.answer_override is not None
    assert "| Archivo | Nombre de Propietario Principal | Nombre Beneficiario |" in resolution.answer_override
    assert "| ESTAN041_ID_14010 | MINISTERIO DE BIENES NACIONALES | MINISTERIO DE BIENES NACIONALES |" in resolution.answer_override
    assert "| ZU163_ID_3630 | ATC SITIOS CHILE S.A. | MINISTERIO DE BIENES NACIONALES |" in resolution.answer_override
    assert "hay 2 beneficiarios distintos" not in resolution.answer_override


def test_question_fact_resolver_answers_contract_state_counts_from_global_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 201,
                    "archive_slug": "A_ID_1",
                    "metadata_json": json.dumps(
                        {"file": "A_ID_1", "fields": {"Estado Contrato": "Vigente"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 202,
                    "archive_slug": "B_ID_2",
                    "metadata_json": json.dumps(
                        {"file": "B_ID_2", "fields": {"Estado Contrato": "Vigente"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 203,
                    "archive_slug": "C_ID_3",
                    "metadata_json": json.dumps(
                        {"file": "C_ID_3", "fields": {"Estado Contrato": "Vencido"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 204,
                    "archive_slug": "D_ID_4",
                    "metadata_json": json.dumps(
                        {"file": "D_ID_4", "fields": {"Estado Contrato": "Terminado"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 205,
                    "archive_slug": "E_ID_5",
                    "metadata_json": json.dumps(
                        {"file": "E_ID_5", "fields": {"Estado Contrato": ""}},
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Cuantos contratos estan vigentes, vencidos o terminados? Reporta tambien los sin estado.",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override == (
        "Según la metadata cargada: vigentes=2, vencidos=1, terminados=1, sin estado=1."
    )
    assert resolution.facts_used_count == 5


def test_question_fact_resolver_answers_top_document_versions_from_inventory() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [],
            file_rows=[
                {"file_id": 301, "archive_slug": "AT565_ID_3820"},
                {"file_id": 302, "archive_slug": "AT565_ID_3820"},
                {"file_id": 303, "archive_slug": "AT565_ID_3820"},
                {"file_id": 304, "archive_slug": "RM797_ID_5515"},
                {"file_id": 305, "archive_slug": "RM797_ID_5515"},
                {"file_id": 306, "archive_slug": "AI041_ID_49"},
            ],
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Que contratos tienen mas versiones documentales o PDFs asociados?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "AT565_ID_3820 (3 PDFs)" in resolution.answer_override
    assert "RM797_ID_5515 (2 PDFs)" in resolution.answer_override
    assert "AI041_ID_49 (1 PDFs)" in resolution.answer_override
    assert resolution.facts_used_count == 6


def test_question_fact_resolver_answers_entel_figure_count_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 401,
                    "archive_slug": "A_ID_1",
                    "metadata_json": json.dumps(
                        {"file": "A_ID_1", "fields": {"Figura Legal": "ENTEL PCS", "Estado Contrato": "Vigente"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 402,
                    "archive_slug": "B_ID_2",
                    "metadata_json": json.dumps(
                        {"file": "B_ID_2", "fields": {"Figura Legal": "ENTEL PCS", "Estado Contrato": "Vigente"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 403,
                    "archive_slug": "C_ID_3",
                    "metadata_json": json.dumps(
                        {"file": "C_ID_3", "fields": {"Figura Legal": "ENTEL PCS", "Estado Contrato": "Terminado"}},
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 404,
                    "archive_slug": "D_ID_4",
                    "metadata_json": json.dumps(
                        {"file": "D_ID_4", "fields": {"Figura Legal": "ENTEL S.A.", "Estado Contrato": "Vigente"}},
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Cuantos contratos vigentes fueron firmados por ENTEL PCS?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override == (
        "Se identificaron 2 contratos vigentes firmados por ENTEL PCS según la metadata cargada."
    )
    assert resolution.facts_used_count == 2


def test_question_fact_resolver_answers_generic_site_count_filtered_by_region() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Codigo de Sitio": "RM797",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Codigo de Sitio": "RM797",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 103,
                    "archive_slug": "SA561_ID_2198",
                    "metadata_json": json.dumps(
                        {
                            "file": "SA561_ID_2198",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Codigo de Sitio": "SA561",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 104,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Region": "Region de Los Lagos",
                                "Codigo de Sitio": "LA122",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Segun la metadata cuantos sitios hay en la region metropolitana de Santiago?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "hay 2 sitios distintos" in resolution.answer_override
    assert "Region=Region Metropolitana de Santiago" in resolution.answer_override
    assert "RM797" in resolution.answer_override
    assert "SA561" in resolution.answer_override
    assert resolution.facts_used_count == 3
    assert resolution.narrowed_file_ids == [101, 102, 103]


def test_question_fact_resolver_metadata_comparison_fallback_handles_region_aggregate() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Codigo de Sitio": "RM797",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Codigo de Sitio": "RM797",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Segun la metadata cuantos sitios hay en la region metropolitana de Santiago?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "hay 1 sitios distintos" in resolution.answer_override
    assert resolution.facts_used_count == 2


def test_question_fact_resolver_answers_natural_contract_listing_with_metadata_filters() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 111,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 112,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Estado Contrato": "Terminado",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 113,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Region": "Region Metropolitana de Santiago",
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="me puedes indicar los contratos vigentes en la region metropolitana?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "hay 2 contratos" in resolution.answer_override
    assert "Estado Contrato=Vigente" in resolution.answer_override
    assert "Region=Region Metropolitana de Santiago" in resolution.answer_override
    assert "RM797_ID_1668" in resolution.answer_override
    assert "LA122_ID_3979" in resolution.answer_override
    assert "RM797_ID_5515" not in resolution.answer_override
    assert resolution.facts_used_count == 2
    assert resolution.narrowed_file_ids == [111, 113]


def test_question_fact_resolver_answers_dynamic_runtime_metadata_lookup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 501,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Segmento Comercial": "Retail Corporativo",
                                "Responsable Comercial": "Camila Soto",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Para RM797_ID_5515, cual es el Segmento Comercial y el Responsable Comercial?",
        user_id=7,
        file_ids=[501],
    )

    assert resolution.narrowed_file_ids == [501]
    assert resolution.answer_override is not None
    assert "| Segmento Comercial | Retail Corporativo |" in resolution.answer_override
    assert "| Responsable Comercial | Camila Soto |" in resolution.answer_override
    assert resolution.facts_used_count == 2


def test_question_fact_resolver_answers_dynamic_metadata_aggregate_from_runtime_schema() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 601,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Segmento Comercial": "Retail",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 602,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Segmento Comercial": "Industrial",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 603,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Segmento Comercial": "Retail",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Segun la metadata, cuantos segmentos comerciales hay?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "hay 2 valores distintos de Segmento Comercial" in resolution.answer_override
    assert "Retail" in resolution.answer_override
    assert "Industrial" in resolution.answer_override
    assert resolution.facts_used_count == 3
    assert resolution.narrowed_file_ids == [601, 602, 603]


def test_question_fact_resolver_summarizes_patterns_for_requested_dynamic_metadata_fields() -> None:
    rows = []
    for index in range(1, 11):
        rows.append(
            {
                "file_id": 800 + index,
                "archive_slug": f"CASE_ID_{index}",
                "metadata_json": json.dumps(
                    {
                        "file": f"CASE_ID_{index}",
                        "fields": {
                            "Región": "Region Metropolitana de Santiago" if index <= 9 else "Region de Valparaiso",
                            "Comuna": "Santiago" if index <= 8 else "Valparaiso",
                            "Tipo de Sitio": "Macro" if index <= 7 else "Indoor",
                            "Tipo de Contrato": "Arriendo" if index <= 8 else "Compra-Venta",
                        },
                    },
                    ensure_ascii=False,
                ),
            }
        )
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(rows),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Existen patrones relevantes por region, comuna, tipo de sitio o tipo de contrato?",
        user_id=7,
        file_ids=[],
        metadata_mode="metadata_first",
        metadata_fields=["Región", "Comuna", "Tipo de Sitio", "Tipo de Contrato"],
    )

    assert resolution.answer_override is not None
    assert "Patrones relevantes" in resolution.answer_override
    assert "8 expedientes" in resolution.answer_override
    assert "Arriendos" in resolution.answer_override
    assert "Compra-Venta" in resolution.answer_override


def test_question_fact_resolver_answers_dynamic_duplicate_metadata_aggregate_from_runtime_schema() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 701,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Segmento Comercial": "Retail",
                                "Folio Interno": "F-01",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 702,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Segmento Comercial": "Retail",
                                "Folio Interno": "F-02",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 703,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Segmento Comercial": "Industrial",
                                "Folio Interno": "F-03",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="Segun la metadata, que segmentos comerciales tienen mas de un folio interno?",
        user_id=7,
        file_ids=[],
    )

    assert resolution.answer_override is not None
    assert "Segmento Comercial" in resolution.answer_override
    assert "Folio Interno" in resolution.answer_override
    assert "Retail: F-01, F-02" in resolution.answer_override
    assert resolution.facts_used_count == 2
    assert resolution.narrowed_file_ids == [701, 702]


def test_question_fact_resolver_confirms_expected_metadata_value() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Id": 5515,
                                "Forma de Pago": "Deposito",
                                "Estado Contrato": "Terminado",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Filtra por file RM797_ID_5515 y revisa si la Forma de Pago reportada en metadata es Deposito.",
        user_id=7,
        file_ids=[102],
    )

    assert resolution.answer_override == "Si, en la metadata de RM797_ID_5515 Forma de Pago es Deposito."
    assert resolution.facts_used_count == 1


def test_question_fact_resolver_uses_beneficiary_aliases_for_rent_receiver_question() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Nombre Beneficiario": "SERVICIOS DE ESTETICA Y BELLEZA PERFECT NAILS LIMITADA",
                                "Rut Beneficiario del contrato": "76433960-6",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Para RM797_ID_5515, quien recibe la renta y cual es su RUT?",
        user_id=7,
        file_ids=[102],
    )

    assert resolution.answer_override is not None
    assert "| Nombre Beneficiario | SERVICIOS DE ESTETICA Y BELLEZA PERFECT NAILS LIMITADA |" in resolution.answer_override
    assert "| Rut Beneficiario del contrato | 76433960-6 |" in resolution.answer_override
    assert resolution.document_phase_required is False


def test_question_fact_resolver_routes_uncovered_metadata_question_to_document_followup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Renta o Precio Vigente": 442,
                                "Pago Anticipado": False,
                                "Periodo de Pago": "Anual",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Renta o Precio Vigente": 45,
                                "Pago Anticipado": False,
                                "Periodo de Pago": "Mensual",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="¿Hay penalización por pago atrasado de renta? RM797",
        user_id=7,
        file_ids=[101, 102],
    )

    assert resolution.narrowed_file_ids == [101, 102]
    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.answerability_route == "metadata_plus_documents"
    assert "Resolved metadata facts:" in resolution.fact_context_text
    assert "RM797_ID_1668: Renta o Precio Vigente=442" in resolution.fact_context_text
    assert "RM797_ID_5515: Renta o Precio Vigente=45" in resolution.fact_context_text
    assert any("not cover the whole question" in note for note in resolution.confidence_notes)


def test_question_fact_resolver_uses_agnostic_metadata_coverage_for_document_followup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 201,
                    "archive_slug": "OP123_ID_A",
                    "metadata_json": json.dumps(
                        {
                            "file": "OP123_ID_A",
                            "fields": {
                                "Fecha de Entrega": "10/01/2026",
                                "Estado del Proceso": "Aprobado",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="¿Hay justificación por retraso en la entrega? OP123",
        user_id=7,
        file_ids=[201],
    )

    assert resolution.narrowed_file_ids == [201]
    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.answerability_route == "metadata_plus_documents"
    assert "OP123_ID_A: Fecha de Entrega=10/01/2026" in resolution.fact_context_text


def test_resolve_facts_keeps_retrieval_open_when_metadata_needs_documents() -> None:
    class _FactResolver:
        def resolve(self, **kwargs: object) -> FactResolution:
            del kwargs
            return FactResolution(
                narrowed_file_ids=[201],
                fact_context_text="Resolved metadata facts:\nOP123_ID_A: Fecha de Entrega=10/01/2026",
                answer_override=None,
                facts_used_count=1,
                metadata_phase_used=True,
                resolved_archive_slugs=["OP123_ID_A"],
                resolved_metadata_fields=["Fecha de Entrega"],
                document_phase_required=True,
                answerability_route="metadata_plus_documents",
            )

    nodes = QAGraphNodes(
        intent_router=object(),
        casual_responder=object(),
        supervisor=object(),
        scope_resolver=object(),
        question_classifier=object(),
        fact_resolver=_FactResolver(),
        retrieval_tool=object(),
        analysis_agent=object(),
        hybrid_answer_tool=object(),
        page_vision_tool=object(),
        repository=object(),
    )

    patch = nodes.resolve_facts(
        {
            "question": "¿Hay justificación por retraso en la entrega? OP123",
            "original_question": "¿Hay justificación por retraso en la entrega? OP123",
            "question_class": "metadata_comparison",
            "user_id": 7,
            "file_ids": [201],
            "top_k": 5,
        }
    )

    assert patch["skip_retrieval"] is False
    assert patch["answer_override"] is None
    assert patch["answerability_route"] == "metadata_plus_documents"
    assert patch["retrieval_route"] == ""


def test_retrieval_question_expands_metadata_plus_documents_generically() -> None:
    expanded = _build_retrieval_question(
        question="¿Hay justificación por retraso en la entrega? OP123",
        answerability_route="metadata_plus_documents",
    )

    assert "justificación por retraso en la entrega" in expanded
    assert "Busqueda documental ampliada" in expanded
    assert "equivalentes" in expanded
    assert "condiciones" in expanded
    assert "consecuencias" in expanded
    assert "ausencia de informacion relevante" in expanded


def test_retrieval_question_does_not_expand_structured_only_metadata_queries() -> None:
    question = "Para RM797_ID_5515, quien recibe la renta y cual es su RUT?"

    assert _build_retrieval_question(
        question=question,
        answerability_route="structured_only",
    ) == question


def test_retrieve_candidates_boosts_metadata_plus_documents_coverage_generically() -> None:
    class _Plan:
        top_k = 5
        strategy = "fast-grounded"
        selected_provider = "fake"

    class _Supervisor:
        def __init__(self) -> None:
            self.question = ""

        def create_plan(self, **kwargs: object) -> _Plan:
            self.question = str(kwargs["question"])
            return _Plan()

    class _RetrievalTool:
        def __init__(self) -> None:
            self.kwargs: dict[str, object] = {}

        def retrieve(self, **kwargs: object) -> RetrievalResult:
            self.kwargs = dict(kwargs)
            return RetrievalResult(evidence=[], telemetry={"retrieval_route": "scoped_semantic"})

    supervisor = _Supervisor()
    retrieval_tool = _RetrievalTool()
    nodes = QAGraphNodes(
        intent_router=object(),
        casual_responder=object(),
        supervisor=supervisor,
        scope_resolver=object(),
        question_classifier=object(),
        fact_resolver=object(),
        retrieval_tool=retrieval_tool,
        analysis_agent=object(),
        hybrid_answer_tool=object(),
        page_vision_tool=object(),
        repository=object(),
    )

    patch = nodes.retrieve_candidates(
        {
            "question": "¿Hay justificación por retraso en la entrega? OP123",
            "top_k": 5,
            "candidate_k": 20,
            "min_pages_per_selected_doc": 0,
            "summary_mode": "default",
            "question_class": "metadata_comparison",
            "user_id": 7,
            "file_ids": [201, 202],
            "resolved_archive_slugs": ["OP123_ID_A", "OP123_ID_B"],
            "answerability_route": "metadata_plus_documents",
            "scope_origin": "metadata",
        }
    )

    assert "Busqueda documental ampliada" in supervisor.question
    assert "justificación por retraso en la entrega" in supervisor.question
    assert retrieval_tool.kwargs["candidate_k"] == 80
    assert retrieval_tool.kwargs["min_pages_per_selected_doc"] == 2
    assert retrieval_tool.kwargs["summary_mode"] == "per_document"
    assert patch["candidate_k"] == 80
    assert patch["min_pages_per_selected_doc"] == 2
    assert patch["summary_mode"] == "per_document"
    assert any("generic per-document retrieval coverage" in note for note in patch["confidence_notes"])


def test_question_fact_resolver_dedupes_repeated_archive_metadata_rows() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 7,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 8,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Filtra por file RM797_ID_5515 y revisa la Forma de Pago reportada en metadata.",
        user_id=7,
        file_ids=[7, 8],
    )

    assert resolution.narrowed_file_ids == [7]
    assert resolution.answer_override == (
        "En la metadata de RM797_ID_5515:\n\n"
        "| Campo | Valor |\n"
        "| --- | --- |\n"
        "| Forma de Pago | Deposito |"
    )
    assert resolution.facts_used_count == 1


def test_question_fact_resolver_compares_metadata_between_files() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Forma de Pago": "Transferencia Electronica",
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Forma de Pago": "Deposito",
                                "Estado Contrato": "Terminado",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Compara la metadata de los archivos RM797_ID_1668 y RM797_ID_5515 "
            "para Forma de Pago y Estado Contrato."
        ),
        user_id=7,
        file_ids=[101, 102],
    )

    assert resolution.narrowed_file_ids == [101, 102]
    assert resolution.answer_override is not None
    assert "| Archivo | Estado Contrato | Forma de Pago |" in resolution.answer_override
    assert "| RM797_ID_1668 | Vigente | Transferencia Electronica |" in resolution.answer_override
    assert "| RM797_ID_5515 | Terminado | Deposito |" in resolution.answer_override
    assert resolution.facts_used_count == 4


def test_question_fact_resolver_routes_interpretive_archive_comparison_to_document_followup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 92,
                    "archive_slug": "AT565_ID_3820",
                    "metadata_json": json.dumps(
                        {
                            "file": "AT565_ID_3820",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Estado Contrato": "desconocido",
                                "Fecha de Término del Contrato": "10/12/2025",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Estado Contrato": "Terminado",
                                "Fecha de Término del Contrato": "22/07/2025",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Compara RM797_ID_1668 y RM797_ID_5515 en Estado Contrato y Fecha de Término del Contrato; "
            "indica si ambos casos parecen vigentes o si hay diferencias claras entre los dos folios."
        ),
        user_id=7,
        file_ids=[91, 92, 101, 102],
    )

    assert resolution.narrowed_file_ids == [101, 102]
    assert resolution.answer_override is None
    assert "Resolved metadata facts:" in resolution.fact_context_text
    assert "RM797_ID_1668:" in resolution.fact_context_text
    assert "Estado Contrato=desconocido" in resolution.fact_context_text
    assert "RM797_ID_5515:" in resolution.fact_context_text
    assert "Estado Contrato=Terminado" in resolution.fact_context_text
    assert "AI041_ID_49" not in resolution.fact_context_text
    assert "AT565_ID_3820" not in resolution.fact_context_text
    assert (
        "Metadata rows matched, but the comparative question requires documentary grounding before drawing conclusions."
        in resolution.confidence_notes
    )
    assert "Archive metadata enriched the structured context for retrieval." in resolution.confidence_notes


def test_question_fact_resolver_routes_metadata_vs_documents_question_to_document_followup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 301,
                    "archive_slug": "TSM10_ID_20441",
                    "metadata_json": json.dumps(
                        {
                            "file": "TSM10_ID_20441",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Revision Final": "Cifrado",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="La metadata coincide con lo que dicen los documentos contractuales vigentes?",
        user_id=7,
        file_ids=[301],
        metadata_fields=["Estado Contrato", "Revision Final"],
    )

    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.narrowed_file_ids == [301]
    assert "Resolved metadata facts:" in resolution.fact_context_text
    assert "TSM10_ID_20441: Estado Contrato=Vigente" in resolution.fact_context_text


def test_question_fact_resolver_uses_docling_quality_for_metadata_document_validation_when_available() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 301,
                    "archive_slug": "CASE_ID_1",
                    "metadata_json": json.dumps(
                        {
                            "file": "CASE_ID_1",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Revision Final": "REVISADO OK",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ],
            page_quality_rows=[
                {
                    "file_id": 301,
                    "archive_slug": "CASE_ID_1",
                    "file_name": "CASE_CONTRATO.pdf",
                    "status": "failed",
                    "file_page_count": 16,
                    "indexed_pages_count": 0,
                    "encrypted_or_unreadable_pages_count": 1,
                    "avg_ocr_confidence": 0.0,
                    "avg_text_quality": 0.0,
                    "ocr_methods": ["docling_rapidocr"],
                    "visual_flags": ["encrypted_pdf"],
                }
            ],
        ),
    )

    resolution = resolver.resolve(
        question_class="analytics",
        question="La metadata coincide con lo que dicen los documentos contractuales vigentes?",
        user_id=7,
        file_ids=[301],
        metadata_fields=["Estado Contrato", "Revision Final"],
    )

    assert resolution.answer_override is not None
    assert "consistencia no validable aun" in resolution.answer_override
    assert "CASE_CONTRATO.pdf" in resolution.answer_override
    assert "PDF cifrado" in resolution.answer_override
    assert resolution.metadata_only_reason == "docling_quality_review"


def test_question_fact_resolver_prioritizes_docling_quality_for_human_review() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 301,
                    "archive_slug": "TSM10_ID_20441",
                    "metadata_json": json.dumps(
                        {
                            "file": "TSM10_ID_20441",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Revisión Final": "REVISADO OK",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ],
            file_rows=[
                {
                    "file_id": 301,
                    "archive_slug": "TSM10_ID_20441",
                    "file_input_file_name": "TSM10_CONTRATO.pdf",
                    "file_code": "TSM10",
                    "file_state": 3,
                    "file_page_count": 1,
                }
            ],
            page_quality_rows=[
                {
                    "file_id": 301,
                    "archive_slug": "TSM10_ID_20441",
                    "file_name": "TSM10_CONTRATO.pdf",
                    "status": "completed",
                    "file_page_count": 1,
                    "indexed_pages_count": 1,
                    "encrypted_or_unreadable_pages_count": 1,
                    "low_ocr_pages_count": 1,
                    "blank_pages_count": 0,
                    "avg_ocr_confidence": 0.41,
                    "min_ocr_confidence": 0.41,
                    "avg_text_quality": 0.32,
                    "ocr_methods": ["docling_rapidocr"],
                    "visual_flags": ["low_ocr_confidence", "encrypted_pdf"],
                }
            ],
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Que campos presentan ausencia, ambiguedad o contradiccion y requieren revision humana?",
        user_id=7,
        file_ids=[301],
        metadata_mode="metadata_first",
        metadata_fields=["Estado Contrato", "Revisión Final"],
    )

    assert resolution.answer_override is not None
    assert "requiere revision humana" in resolution.answer_override
    assert "TSM10_CONTRATO.pdf" in resolution.answer_override
    assert "cifrado" in resolution.answer_override
    assert "REVISADO OK" in resolution.answer_override
    assert resolution.metadata_phase_used is True
    assert resolution.metadata_only_reason == "docling_quality_review"


def test_question_fact_resolver_does_not_short_circuit_ocr_content_request_to_quality_review() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 301,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Codigo de Sitio": "AI041",
                                "Direccion": "Av. Costanera Sur 2760",
                                "Renta o Precio Vigente": "14 UF mensuales",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ],
            file_rows=[
                {
                    "file_id": 301,
                    "archive_slug": "AI041_ID_49",
                    "file_input_file_name": "AI041_CONTRATO.pdf",
                    "file_code": "AI041",
                    "file_state": 3,
                    "file_page_count": 18,
                }
            ],
            page_quality_rows=[
                {
                    "file_id": 301,
                    "archive_slug": "AI041_ID_49",
                    "file_name": "AI041_CONTRATO.pdf",
                    "status": "completed",
                    "file_page_count": 18,
                    "indexed_pages_count": 18,
                    "encrypted_or_unreadable_pages_count": 0,
                    "low_ocr_pages_count": 1,
                    "blank_pages_count": 0,
                    "avg_ocr_confidence": 0.61,
                    "avg_text_quality": 0.45,
                    "ocr_methods": ["docling_rapidocr"],
                    "visual_flags": ["low_ocr_confidence"],
                }
            ],
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usar @metadata y /file:AI041_ID_49. Segun el OCR del documento, resume de que trata "
            "el contrato y menciona las partes principales, la direccion o sitio y la renta si aparece."
        ),
        user_id=7,
        file_ids=[301],
        metadata_mode="metadata_first",
    )

    assert resolution.answer_override is None
    assert resolution.document_phase_required is True
    assert resolution.metadata_only_reason == ""
    assert resolution.narrowed_file_ids == [301]
    assert "Archive metadata context:" in resolution.fact_context_text
    assert "Renta o Precio Vigente=14 UF mensuales" in resolution.fact_context_text


def test_question_fact_resolver_counts_docling_encrypted_pdfs_for_review_priority() -> None:
    metadata_rows = []
    file_rows = []
    page_quality_rows = []
    for index, archive_slug in enumerate(["ZB352_ID_2668", "RM797_ID_5515", "TSM10_ID_20441"], start=1):
        metadata_rows.append(
            {
                "file_id": index,
                "archive_slug": archive_slug,
                "metadata_json": json.dumps(
                    {
                        "file": archive_slug,
                        "fields": {
                            "Estado Contrato": "Vigente",
                            "Revisión Final": "REVISADO OK",
                        },
                    },
                    ensure_ascii=False,
                ),
            }
        )
        file_rows.append(
            {
                "file_id": index,
                "archive_slug": archive_slug,
                "file_input_file_name": f"{archive_slug}_CONTRATO.pdf",
                "file_code": archive_slug.split("_", 1)[0],
                "file_state": 3,
                "file_page_count": 1,
            }
        )
        page_quality_rows.append(
            {
                "file_id": index,
                "archive_slug": archive_slug,
                "file_name": f"{archive_slug}_CONTRATO.pdf",
                "status": "completed",
                "file_page_count": 1,
                "indexed_pages_count": 1,
                "encrypted_or_unreadable_pages_count": 1,
                "low_ocr_pages_count": 1,
                "blank_pages_count": 0,
                "avg_ocr_confidence": 0.25,
                "avg_text_quality": 0.2,
                "ocr_methods": ["docling_rapidocr"],
                "visual_flags": ["encrypted_pdf"],
            }
        )
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            metadata_rows,
            file_rows=file_rows,
            page_quality_rows=page_quality_rows,
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Que expedientes deberian priorizarse para revision humana por OCR, cifrado o inconsistencias?",
        user_id=7,
        file_ids=[1, 2, 3],
        metadata_mode="metadata_first",
        metadata_fields=["Revisión Final", "Estado Contrato"],
    )

    assert resolution.answer_override is not None
    assert "3 PDF cifrados" in resolution.answer_override
    assert "1 de 1 PDF cifrado" in resolution.answer_override
    assert "consistencia no validable aun" in resolution.answer_override


def test_question_fact_resolver_extracts_date_field_when_question_text_is_garbled() -> None:
    metadata_rows = [
        ArchiveMetadataEntry(
            file_id=101,
            archive_slug="RM797_ID_1668",
            fields={
                "Estado Contrato": "desconocido",
                "Fecha de Término del Contrato": "10/12/2025",
            },
        ),
        ArchiveMetadataEntry(
            file_id=102,
            archive_slug="RM797_ID_5515",
            fields={
                "Estado Contrato": "Terminado",
                "Fecha de Término del Contrato": "22/07/2025",
            },
        ),
    ]

    requested_fields = QuestionFactResolver._extract_requested_metadata_fields(
        question=(
            "Compara RM797_ID_1668 y RM797_ID_5515 en Estado Contrato y "
            "Fecha de T?rmino del Contrato."
        ),
        metadata_rows=metadata_rows,
    )

    assert "Estado Contrato" in requested_fields
    assert "Fecha de Término del Contrato" in requested_fields


def test_question_fact_resolver_routes_missing_metadata_comparison_to_document_followup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 101,
                    "archive_slug": "RM797_ID_1668",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_1668",
                            "fields": {
                                "Fecha de TÃ©rmino del Contrato": "10/12/2025",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 102,
                    "archive_slug": "RM797_ID_5515",
                    "metadata_json": json.dumps(
                        {
                            "file": "RM797_ID_5515",
                            "fields": {
                                "Estado Contrato": "Terminado",
                                "Fecha de TÃ©rmino del Contrato": "22/07/2025",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question="Compara RM797_ID_1668 y RM797_ID_5515 en Estado Contrato y Fecha de TÃ©rmino del Contrato.",
        user_id=7,
        file_ids=[101, 102],
    )

    assert resolution.narrowed_file_ids == [101, 102]
    assert resolution.answer_override is None
    assert "Resolved metadata facts:" in resolution.fact_context_text
    assert "RM797_ID_1668: Fecha de TÃ©rmino del Contrato=10/12/2025" in resolution.fact_context_text
    assert "RM797_ID_5515:" in resolution.fact_context_text
    assert "Estado Contrato=Terminado" in resolution.fact_context_text
    assert (
        "Metadata rows matched partially, but missing values require documentary evidence before concluding the comparison."
        in resolution.confidence_notes
    )
    assert "Archive metadata enriched the structured context for retrieval." in resolution.confidence_notes


def test_question_fact_resolver_uses_aliases_for_compound_single_archive_metadata_question() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Nombre de Propietario Principal": "FILADELFIA DE LA PENA ECHAVEGUREN",
                                "Nombre Beneficiario": "ATC SITIOS CHILE S.A.",
                                "Forma de Pago": "Vale Vista",
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usando metadata del archivo AI041_ID_49 valida propietario principal, "
            "beneficiario actual, forma de pago y si el contrato sigue vigente."
        ),
        user_id=7,
        file_ids=[91],
    )

    assert resolution.narrowed_file_ids == [91]
    assert resolution.answer_override is not None
    assert "| Nombre de Propietario Principal | FILADELFIA DE LA PENA ECHAVEGUREN |" in resolution.answer_override
    assert "| Nombre Beneficiario | ATC SITIOS CHILE S.A. |" in resolution.answer_override
    assert "| Forma de Pago | Vale Vista |" in resolution.answer_override
    assert "| Estado Contrato | Vigente |" in resolution.answer_override
    assert resolution.facts_used_count == 4


def test_question_fact_resolver_validates_multiple_expected_values_for_single_archive() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Pago Anticipado": True,
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usando metadata de AI041_ID_49, valida si Estado Contrato es Vigente, "
            "Pago Anticipado es SI y Forma de Pago es Vale Vista."
        ),
        user_id=7,
        file_ids=[91],
    )

    assert resolution.answer_override is not None
    assert "| Estado Contrato | Vigente | coincide |" in resolution.answer_override
    assert "| Pago Anticipado | SI | coincide |" in resolution.answer_override
    assert "| Forma de Pago | Vale Vista | coincide |" in resolution.answer_override
    assert resolution.facts_used_count == 3


def test_question_fact_resolver_keeps_metadata_as_context_for_mixed_document_question() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usando metadata y documentos de AI041_ID_49, valida si el Estado Contrato es Vigente "
            "y respaldalo con AI041.pdf."
        ),
        user_id=7,
        file_ids=[91],
    )

    assert resolution.narrowed_file_ids == [91]
    assert resolution.answer_override is None
    assert "Resolved metadata facts:" in resolution.fact_context_text
    assert "AI041_ID_49: Estado Contrato=Vigente" in resolution.fact_context_text


def test_question_fact_resolver_raises_when_metadata_is_requested_but_unavailable() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository([]),
    )

    with pytest.raises(ScopeResolutionError) as exc_info:
        resolver.resolve(
            question_class="extractive",
            question="Consulta con metadata sobre el propietario",
            user_id=7,
            file_ids=[91],
            metadata_mode="metadata_first",
        )

    assert exc_info.value.status_code == 404
    assert "No metadata is available" in str(exc_info.value)


def test_question_fact_resolver_raises_when_structured_metadata_field_is_unknown() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    with pytest.raises(ScopeResolutionError) as exc_info:
        resolver.resolve(
            question_class="extractive",
            question="Consulta el valor solicitado",
            user_id=7,
            file_ids=[91],
            metadata_fields=["Columna Inexistente"],
        )

    assert exc_info.value.status_code == 404
    assert "Columna Inexistente" in str(exc_info.value)


def test_question_fact_resolver_uses_structured_metadata_fields_for_metadata_first_lookup() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ],
            file_rows=[
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "file_input_file_name": "AI041.pdf",
                    "file_state": 3,
                    "file_page_count": 18,
                },
                {
                    "file_id": 92,
                    "archive_slug": "AI041_ID_49",
                    "file_input_file_name": "AI041_anexo.pdf",
                    "file_state": 3,
                    "file_page_count": 6,
                },
            ],
        ),
    )

    resolution = resolver.resolve(
        question_class="extractive",
        question="Dime el valor actual",
        user_id=7,
        file_ids=[91],
        metadata_mode="metadata_first",
        archive_slugs=["AI041_ID_49"],
        metadata_fields=["Estado Contrato"],
    )

    assert resolution.answer_override == (
        "En la metadata de AI041_ID_49:\n\n"
        "| Campo | Valor |\n"
        "| --- | --- |\n"
        "| Estado Contrato | Vigente |"
    )
    assert "Documentos asociados" not in (resolution.answer_override or "")
    assert resolution.metadata_phase_used is True
    assert resolution.resolved_archive_slugs == ["AI041_ID_49"]
    assert resolution.resolved_metadata_fields == ["Estado Contrato"]
    assert resolution.metadata_only_reason == "metadata_fields_sufficient"


def test_question_fact_resolver_keeps_only_matching_archive_context_for_broad_manual_scope() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Vale Vista",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 92,
                    "archive_slug": "AT565_ID_3820",
                    "metadata_json": json.dumps(
                        {
                            "file": "AT565_ID_3820",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usando metadata y documentos de AT565_ID_3820, valida si el Estado Contrato es Vigente "
            "y respaldalo con AT565.PDF."
        ),
        user_id=7,
        file_ids=[91, 92],
    )

    assert resolution.narrowed_file_ids == [92]
    assert resolution.answer_override is None
    assert "AT565_ID_3820: Estado Contrato=Vigente" in resolution.fact_context_text
    assert "AI041_ID_49" not in resolution.fact_context_text


def test_question_fact_resolver_expands_document_evidence_scope_to_all_matching_archive_files() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 91,
                    "archive_slug": "AT565_ID_3820",
                    "metadata_json": json.dumps(
                        {
                            "file": "AT565_ID_3820",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 92,
                    "archive_slug": "AT565_ID_3820",
                    "metadata_json": json.dumps(
                        {
                            "file": "AT565_ID_3820",
                            "fields": {
                                "Estado Contrato": "Vigente",
                                "Forma de Pago": "Deposito",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 93,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Estado Contrato": "Vigente",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Usando metadata y documentos de AT565_ID_3820, valida si el Estado Contrato es Vigente "
            "y respaldalo con AT565.PDF."
        ),
        user_id=7,
        file_ids=[91, 92, 93],
    )

    assert resolution.narrowed_file_ids == [91, 92]
    assert resolution.answer_override is None
    assert "AT565_ID_3820: Estado Contrato=Vigente" in resolution.fact_context_text


def test_question_fact_resolver_flags_missing_metadata_in_single_field_comparison() -> None:
    answer = QuestionFactResolver._build_metadata_answer(
        question="Compara RM797_ID_1668 y RM797_ID_5515 en Estado Contrato.",
        metadata_rows=[
            ArchiveMetadataEntry(
                file_id=101,
                archive_slug="RM797_ID_1668",
                fields={"Fecha de Término del Contrato": "10/12/2025"},
            ),
            ArchiveMetadataEntry(
                file_id=102,
                archive_slug="RM797_ID_5515",
                fields={"Estado Contrato": "Terminado"},
            ),
        ],
        requested_fields=["Estado Contrato"],
        compare_requested=True,
    )

    assert answer == (
        "Estado Contrato por archivo:\n\n"
        "| Archivo | Estado Contrato |\n"
        "| --- | --- |\n"
        "| RM797_ID_5515 | Terminado |\n\n"
        "Falta metadata para: RM797_ID_1668."
    )


def test_question_fact_resolver_formats_single_field_listing_as_markdown_table() -> None:
    answer = QuestionFactResolver._build_metadata_answer(
        question="Lista la",
        metadata_rows=[
            ArchiveMetadataEntry(
                file_id=101,
                archive_slug="AI041_ID_49",
                fields={"Renta o Precio Vigente": "420"},
            ),
            ArchiveMetadataEntry(
                file_id=102,
                archive_slug="RM797_ID_1668",
                fields={"Renta o Precio Vigente": "442"},
            ),
        ],
        requested_fields=["Renta o Precio Vigente"],
        compare_requested=False,
    )

    assert answer == (
        "Renta o Precio Vigente por archivo:\n\n"
        "| Archivo | Renta o Precio Vigente |\n"
        "| --- | --- |\n"
        "| AI041_ID_49 | 420 |\n"
        "| RM797_ID_1668 | 442 |"
    )
    assert "; " not in answer


def test_question_fact_resolver_derives_monthly_equivalent_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 401,
                    "archive_slug": "AI041_ID_49",
                    "metadata_json": json.dumps(
                        {
                            "file": "AI041_ID_49",
                            "fields": {
                                "Renta o Precio Vigente": "420",
                                "Periodo de Pago": "Cada 4 años",
                                "Tipo de Moneda": "UF",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="extractive",
        question="Cual es el canon mensual equivalente del contrato vigente segun su periodicidad de pago?",
        user_id=7,
        file_ids=[401],
        metadata_fields=["Renta o Precio Vigente", "Periodo de Pago", "Tipo de Moneda"],
    )

    assert resolution.answer_override is not None
    assert "420 UF cada 4 años" in resolution.answer_override
    assert "8,75 UF por mes" in resolution.answer_override


def test_question_fact_resolver_derives_total_canon_for_initial_term_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 402,
                    "archive_slug": "FG459_ID_30575",
                    "metadata_json": json.dumps(
                        {
                            "file": "FG459_ID_30575",
                            "fields": {
                                "Renta o Precio Vigente": "14",
                                "Periodo de Pago": "Mensual",
                                "Tipo de Moneda": "UF",
                                "Duración Inicial del Contrato": "01A-06M-00D",
                                "Fecha de Inicio de Vigencia del Contrato": "14/11/2024",
                                "Fecha de Término del Contrato": "14/05/2026",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="extractive",
        question="Cual es el total estimado de pagos por canon durante la vigencia inicial del contrato?",
        user_id=7,
        file_ids=[402],
        metadata_fields=[
            "Renta o Precio Vigente",
            "Periodo de Pago",
            "Duración Inicial del Contrato",
            "Fecha de Inicio de Vigencia del Contrato",
            "Fecha de Término del Contrato",
        ],
    )

    assert resolution.answer_override is not None
    assert "14 UF mensuales" in resolution.answer_override
    assert "1 año y 6 meses" in resolution.answer_override
    assert "252 UF" in resolution.answer_override


def test_question_fact_resolver_derives_remaining_term_and_key_dates_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 403,
                    "archive_slug": "FG459_ID_30575",
                    "metadata_json": json.dumps(
                        {
                            "file": "FG459_ID_30575",
                            "fields": {
                                "Fecha de Término del Contrato": "14/05/2026",
                                "Fecha de Aviso de Término del Contrato": "14/04/2026",
                                "Prórroga Automática": "NO",
                                "Periodo Prórroga Automática": "01A-00M-00D",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="exhaustive_synthesis",
        question="Cuanto tiempo resta para el vencimiento o renovacion del contrato y que hitos deben vigilarse?",
        user_id=7,
        file_ids=[403],
        metadata_fields=[
            "Fecha de Término del Contrato",
            "Fecha de Aviso de Término del Contrato",
            "Prórroga Automática",
            "Periodo Prórroga Automática",
        ],
        reference_date=date(2026, 4, 25),
    )

    assert resolution.answer_override is not None
    assert "19 dias" in resolution.answer_override
    assert "14/05/2026" in resolution.answer_override
    assert "14/04/2026" in resolution.answer_override
    assert "Prórroga Automática: NO" in resolution.answer_override


def test_question_fact_resolver_derives_contract_term_summary_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 404,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Fecha de Inicio de Vigencia del Contrato": "01/08/2025",
                                "Fecha de TÃ©rmino del Contrato": "01/08/2027",
                                "DuraciÃ³n Inicial del Contrato": "02A-00M-00D",
                                "PrÃ³rroga AutomÃ¡tica": "SI",
                                "Periodo PrÃ³rroga AutomÃ¡tica": "02A-00M-00D",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="extractive",
        question="Cual es el plazo contractual vigente y cual es la fecha de termino del contrato?",
        user_id=7,
        file_ids=[404],
        metadata_fields=[
            "Fecha de Inicio de Vigencia del Contrato",
            "Fecha de TÃ©rmino del Contrato",
            "DuraciÃ³n Inicial del Contrato",
            "PrÃ³rroga AutomÃ¡tica",
            "Periodo PrÃ³rroga AutomÃ¡tica",
        ],
    )

    assert resolution.answer_override is not None
    assert "2 años" in resolution.answer_override
    assert "01/08/2027" in resolution.answer_override
    assert "Prórroga Automática: SI" in resolution.answer_override


def test_question_fact_resolver_derives_next_renewal_window_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 405,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "PrÃ³rroga AutomÃ¡tica": "SI",
                                "Periodo PrÃ³rroga AutomÃ¡tica": "02A-00M-00D",
                                "Fecha de TÃ©rmino del Contrato": "01/08/2027",
                                "Fecha de Aviso de TÃ©rmino del Contrato": "01/02/2027",
                            },
                        },
                        ensure_ascii=False,
                    ),
                }
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="extractive",
        question="Si existe prorroga automatica, cual seria el siguiente vencimiento o ventana de salida?",
        user_id=7,
        file_ids=[405],
        metadata_fields=[
            "PrÃ³rroga AutomÃ¡tica",
            "Periodo PrÃ³rroga AutomÃ¡tica",
            "Fecha de TÃ©rmino del Contrato",
            "Fecha de Aviso de TÃ©rmino del Contrato",
        ],
    )

    assert resolution.answer_override is not None
    assert "01/08/2029" in resolution.answer_override
    assert "2 años" in resolution.answer_override


def test_question_fact_resolver_derives_multi_row_rent_change_summary_from_metadata() -> None:
    resolver = QuestionFactResolver(
        repository=object(),
        file_repository=_FakeArchiveMetadataFileRepository(
            [
                {
                    "file_id": 406,
                    "archive_slug": "LA122_ID_18467",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_18467",
                            "fields": {
                                "Codigo de Sitio": "LA122",
                                "Estado Contrato": "Vencido",
                                "Renta o Precio Vigente": "43.9",
                                "Tipo de Moneda": "UF",
                                "Periodo de Pago": "Mensual",
                                "Fecha de TÃ©rmino del Contrato": "01/01/2022",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
                {
                    "file_id": 407,
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps(
                        {
                            "file": "LA122_ID_3979",
                            "fields": {
                                "Codigo de Sitio": "LA122",
                                "Estado Contrato": "Vigente",
                                "Renta o Precio Vigente": "504",
                                "Tipo de Moneda": "UF",
                                "Periodo de Pago": "Anual",
                                "Fecha de TÃ©rmino del Contrato": "01/08/2027",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        ),
    )

    resolution = resolver.resolve(
        question_class="metadata_comparison",
        question=(
            "Como cambian la renta, plazo y estado contractual entre contratos del mismo propietario, "
            "beneficiario o codigo de sitio?"
        ),
        user_id=7,
        file_ids=[406, 407],
        metadata_fields=[
            "Codigo de Sitio",
            "Estado Contrato",
            "Renta o Precio Vigente",
            "Tipo de Moneda",
            "Periodo de Pago",
            "Fecha de TÃ©rmino del Contrato",
        ],
    )

    assert resolution.answer_override is not None
    assert "504 UF anuales" in resolution.answer_override
    assert "43,9 UF mensuales" in resolution.answer_override


def test_question_fact_resolver_marks_missing_metadata_in_multi_field_comparison() -> None:
    answer = QuestionFactResolver._build_metadata_answer(
        question=(
            "Compara RM797_ID_1668 y RM797_ID_5515 en Estado Contrato y Fecha de Término del Contrato."
        ),
        metadata_rows=[
            ArchiveMetadataEntry(
                file_id=101,
                archive_slug="RM797_ID_1668",
                fields={"Fecha de Término del Contrato": "10/12/2025"},
            ),
            ArchiveMetadataEntry(
                file_id=102,
                archive_slug="RM797_ID_5515",
                fields={
                    "Estado Contrato": "Terminado",
                    "Fecha de Término del Contrato": "22/07/2025",
                },
            ),
        ],
        requested_fields=["Estado Contrato", "Fecha de Término del Contrato"],
        compare_requested=True,
    )

    assert answer == (
        "Comparacion de metadata:\n\n"
        "| Archivo | Estado Contrato | Fecha de Término del Contrato |\n"
        "| --- | --- | --- |\n"
        "| RM797_ID_1668 | sin metadata | 10/12/2025 |\n"
        "| RM797_ID_5515 | Terminado | 22/07/2025 |"
    )


def test_ingestion_override_keeps_document_code_metadata() -> None:
    service = _build_ingestion_service_for_tests()
    detected_metadata = FileMetadata(
        document_code="RM797",
        document_code_source="filename_rule",
    )

    resolved = service._resolve_document_metadata_with_override(
        detected_metadata=detected_metadata,
        metadata_override={
            "document_code": "RM797",
            "document_code_source": "filename_rule",
        },
    )

    assert resolved.document_code == "RM797"
    assert resolved.document_code_source == "filename_rule"


def test_ingestion_without_preview_keeps_detected_document_code_only() -> None:
    service = _build_ingestion_service_for_tests()
    detected_metadata = FileMetadata(
        document_code="RM797",
        document_code_source="filename_rule",
    )

    resolved = service._resolve_document_metadata_with_override(
        detected_metadata=detected_metadata,
        metadata_override=None,
    )

    assert resolved.document_code == "RM797"
    assert resolved.document_code_source == "filename_rule"


def test_metadata_upload_requires_exact_first_file_column(tmp_path: Path) -> None:
    service = _build_metadata_upload_service(tmp_path=tmp_path)
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        "File,Id\nRM797_ID_1668,1668\n",
        encoding="utf-8",
    )

    with pytest.raises(MetadataUploadValidationError, match="first column named `file`"):
        service.parse_csv(csv_path=csv_path)


def test_metadata_upload_rejects_duplicate_files_case_insensitive(tmp_path: Path) -> None:
    service = _build_metadata_upload_service(tmp_path=tmp_path)
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        "file,Id\nRM797_ID_1668.zip,1668\nrm797_id_1668,1669\n",
        encoding="utf-8",
    )

    with pytest.raises(MetadataUploadValidationError, match="Duplicate `file` values"):
        service.upload_csv(
            user_id=7,
            csv_path=csv_path,
            source_file_name="metadata.csv",
        )


def test_metadata_upload_upload_csv_coerces_scalars_and_preserves_dynamic_fields(tmp_path: Path) -> None:
    repository = _FakeMetadataRepository()
    service = _build_metadata_upload_service(
        tmp_path=tmp_path,
        repository=repository,
    )
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        (
            "file,Codigo de Sitio,Id,Activo,Monto,Leading,Observacion\n"
            "RM797_ID_1668.zip,RM797,1668,si,10.5,00123,\n"
            "RM999_ID_1.pdf,RM999,1,no,0,0001,comparar\n"
        ),
        encoding="utf-8",
    )

    result = service.upload_csv(
        user_id=7,
        csv_path=csv_path,
        source_file_name="metadata.csv",
    )

    assert result.columns == [
        "file",
        "Codigo de Sitio",
        "Id",
        "Activo",
        "Monto",
        "Leading",
        "Observacion",
    ]
    assert result.total_rows == 2
    assert result.matched_files == ["RM797_ID_1668"]
    assert result.unmatched_files == ["RM999_ID_1"]
    assert len(repository.created_uploads) == 1
    assert len(repository.replaced_rows) == 1

    stored_rows = repository.replaced_rows[0]["rows"]
    assert len(stored_rows) == 2

    first_row = stored_rows[0]
    assert first_row["file_key"] == "RM797_ID_1668"
    assert first_row["search_text"] == (
        "file: RM797_ID_1668 | Codigo de Sitio: RM797 | Id: 1668 | Activo: True | "
        "Monto: 10.5 | Leading: 00123"
    )
    assert json.loads(first_row["row_json"]) == {
        "file": "RM797_ID_1668",
        "fields": {
            "Codigo de Sitio": "RM797",
            "Id": 1668,
            "Activo": True,
            "Monto": 10.5,
            "Leading": "00123",
            "Observacion": None,
        },
    }

    second_row = stored_rows[1]
    assert json.loads(second_row["row_json"]) == {
        "file": "RM999_ID_1",
        "fields": {
            "Codigo de Sitio": "RM999",
            "Id": 1,
            "Activo": False,
            "Monto": 0,
            "Leading": "0001",
            "Observacion": "comparar",
        },
    }


def test_metadata_upload_replace_csv_preserves_dataset_id_and_replaces_rows(tmp_path: Path) -> None:
    repository = _FakeMetadataRepository()
    service = _build_metadata_upload_service(
        tmp_path=tmp_path,
        repository=repository,
    )
    csv_path = tmp_path / "metadata-v2.csv"
    csv_path.write_text(
        (
            "file,Id,Estado\n"
            "RM797_ID_1668,1668,Activo\n"
            "RM797_ID_5515,5515,Revision\n"
        ),
        encoding="utf-8",
    )

    result = service.replace_csv(
        metadata_upload_id="metadata-123",
        user_id=7,
        csv_path=csv_path,
        source_file_name="metadata-v2.csv",
    )

    assert result.metadata_upload_id == "metadata-123"
    assert result.columns == ["file", "Id", "Estado"]
    assert result.total_rows == 2
    assert repository.updated_uploads[0]["metadata_upload_id"] == "metadata-123"
    assert repository.replaced_rows[0]["metadata_upload_id"] == "metadata-123"
    assert [row["file_key"] for row in repository.replaced_rows[0]["rows"]] == [
        "RM797_ID_1668",
        "RM797_ID_5515",
    ]
    assert repository.refresh_calls == [{"metadata_upload_id": "metadata-123", "user_id": 7}]


def test_archive_metadata_bootstrap_sql_files_exist() -> None:
    sql_dir = Path("apps/backend/db/bootstrap/sql")
    assert (sql_dir / "22_archive_metadata_uploads.sql").exists()
    assert (sql_dir / "23_archive_metadata_upload_rows.sql").exists()
    assert (sql_dir / "24_archive_metadata.sql").exists()


def test_archive_metadata_repository_file_lookup_does_not_use_distinct_with_clob() -> None:
    class _FakeCursor:
        def __init__(self) -> None:
            self.executed_sql = ""

        def execute(self, sql: str, params: dict | None = None) -> None:
            del params
            self.executed_sql = sql

        def fetchall(self) -> list[tuple]:
            return []

        def close(self) -> None:
            return None

    class _FakeConnection:
        def __init__(self, cursor: _FakeCursor) -> None:
            self._cursor = cursor

        def cursor(self) -> _FakeCursor:
            return self._cursor

        def close(self) -> None:
            return None

    fake_cursor = _FakeCursor()
    repository = object.__new__(ArchiveMetadataRepository)
    repository.db_manager = type(
        "_FakeDbManager",
        (),
        {"get_connection": lambda self: _FakeConnection(fake_cursor)},
    )()
    rows = repository.get_archive_metadata_for_file_ids(user_id=7, file_ids=[101])

    assert rows == []
    assert "SELECT DISTINCT" not in _normalize_sql_whitespace(fake_cursor.executed_sql).upper()


def test_archive_metadata_repository_user_lookup_does_not_group_by_metadata_clob() -> None:
    class _FakeCursor:
        def __init__(self) -> None:
            self.executed_sql = ""

        def execute(self, sql: str, params: dict | None = None, **kwargs) -> None:
            del params, kwargs
            self.executed_sql = sql

        def fetchall(self) -> list[tuple]:
            return []

        def close(self) -> None:
            return None

    class _FakeConnection:
        def __init__(self, cursor: _FakeCursor) -> None:
            self._cursor = cursor

        def cursor(self) -> _FakeCursor:
            return self._cursor

        def close(self) -> None:
            return None

    fake_cursor = _FakeCursor()
    repository = object.__new__(ArchiveMetadataRepository)
    repository.db_manager = type(
        "_FakeDbManager",
        (),
        {"get_connection": lambda self: _FakeConnection(fake_cursor)},
    )()

    rows = repository.list_archive_metadata_for_user(user_id=7, include_shared=True)

    normalized_sql = _normalize_sql_whitespace(fake_cursor.executed_sql).upper()
    group_by_fragment = normalized_sql.split("GROUP BY", 1)[1]
    assert rows == []
    assert "AM.METADATA_JSON" not in group_by_fragment
    assert "AM.METADATA_SEARCH_TEXT" not in group_by_fragment


def test_archive_metadata_repository_lists_shared_metadata_uploads_against_visible_files() -> None:
    class _FakeCursor:
        def __init__(self) -> None:
            self.executed_sql = ""
            self.params: dict | None = None

        def execute(self, sql: str, params: dict | None = None, **kwargs) -> None:
            del kwargs
            self.executed_sql = sql
            self.params = params

        def fetchall(self) -> list[tuple]:
            return []

        def close(self) -> None:
            return None

    class _FakeConnection:
        def __init__(self, cursor: _FakeCursor) -> None:
            self._cursor = cursor

        def cursor(self) -> _FakeCursor:
            return self._cursor

        def close(self) -> None:
            return None

    fake_cursor = _FakeCursor()
    repository = object.__new__(ArchiveMetadataRepository)
    repository.db_manager = type(
        "_FakeDbManager",
        (),
        {"get_connection": lambda self: _FakeConnection(fake_cursor)},
    )()

    rows = repository.list_uploads_for_user(user_id=11, include_archived=False, include_shared=True)

    normalized_sql = _normalize_sql_whitespace(fake_cursor.executed_sql).upper()
    assert rows == []
    assert fake_cursor.params == {"user_id": 11}
    assert "U.ACCESS_SCOPE" in normalized_sql
    assert "LOWER(NVL(U.ACCESS_SCOPE, 'PRIVATE')) = 'ALL'" in normalized_sql
    assert "LOWER(NVL(F.ACCESS_SCOPE, 'PRIVATE')) = 'ALL'" in normalized_sql
    assert "F.USER_ID = R.USER_ID" not in normalized_sql


def test_archive_metadata_repository_get_upload_row_can_read_shared_dataset_rows() -> None:
    class _FakeCursor:
        def __init__(self) -> None:
            self.executed_sql = ""

        def execute(self, sql: str, params: dict | None = None, **kwargs) -> None:
            del params, kwargs
            self.executed_sql = sql

        def fetchone(self) -> None:
            return None

        def close(self) -> None:
            return None

    class _FakeConnection:
        def __init__(self, cursor: _FakeCursor) -> None:
            self._cursor = cursor

        def cursor(self) -> _FakeCursor:
            return self._cursor

        def close(self) -> None:
            return None

    fake_cursor = _FakeCursor()
    repository = object.__new__(ArchiveMetadataRepository)
    repository.db_manager = type(
        "_FakeDbManager",
        (),
        {"get_connection": lambda self: _FakeConnection(fake_cursor)},
    )()

    row = repository.get_upload_row(metadata_upload_id="meta-1", user_id=22, file_key="LA122")

    normalized_sql = _normalize_sql_whitespace(fake_cursor.executed_sql).upper()
    assert row is None
    assert "JOIN ARCHIVE_METADATA_UPLOADS U" in normalized_sql
    assert "LOWER(NVL(U.ACCESS_SCOPE, 'PRIVATE')) = 'ALL'" in normalized_sql
    assert "R.USER_ID = :USER_ID" not in normalized_sql


def test_archive_metadata_repository_retries_update_after_unique_constraint_race(monkeypatch: pytest.MonkeyPatch) -> None:
    class _FakeCursor:
        def __init__(self) -> None:
            self.executed_sql: list[str] = []
            self.rowcount = 0

        def execute(self, sql: str, *args, **kwargs) -> None:
            del args, kwargs
            normalized_sql = _normalize_sql_whitespace(sql).upper()
            self.executed_sql.append(normalized_sql)
            if normalized_sql.startswith("MERGE INTO ARCHIVE_METADATA"):
                raise RuntimeError("ORA-00001: unique constraint (APP_DOC.UQ_ARCHIVE_METADATA_SLUG)")
            if normalized_sql.startswith("UPDATE ARCHIVE_METADATA"):
                self.rowcount = 1
                return
            raise AssertionError(f"Unexpected SQL executed: {normalized_sql}")

        def close(self) -> None:
            return None

    class _FakeConnection:
        def __init__(self) -> None:
            self.rollback_calls = 0
            self.commit_calls = 0

        def cursor(self) -> _FakeCursor:
            return fake_cursor

        def rollback(self) -> None:
            self.rollback_calls += 1

        def commit(self) -> None:
            self.commit_calls += 1

        def close(self) -> None:
            return None

    fake_cursor = _FakeCursor()
    fake_connections: list[_FakeConnection] = []

    def _get_connection():
        connection = _FakeConnection()
        fake_connections.append(connection)
        return connection

    import apps.backend.app.repositories.archive_metadata_repository as archive_metadata_module

    monkeypatch.setattr(
        archive_metadata_module,
        "execute_with_oracle_text_repair",
        lambda *, db_manager, operation, candidate_index_names: operation(),
    )

    repository = object.__new__(ArchiveMetadataRepository)
    repository.db_manager = type("_FakeDbManager", (), {"get_connection": staticmethod(_get_connection)})()
    repository.upsert_archive_metadata(
        user_id=7,
        archive_slug="AI041_ID_49",
        metadata_upload_id="upload-123",
        metadata_json='{"file":"AI041_ID_49"}',
        metadata_search_text="file ai041 id 49",
    )

    assert len(fake_connections) == 2
    assert fake_connections[0].rollback_calls == 1
    assert fake_connections[0].commit_calls == 0
    assert fake_connections[1].commit_calls == 1
    assert any(sql.startswith("MERGE INTO ARCHIVE_METADATA") for sql in fake_cursor.executed_sql)
    assert any(sql.startswith("UPDATE ARCHIVE_METADATA") for sql in fake_cursor.executed_sql)


def test_retrieval_metadata_prefilter_keeps_limit_sized_matches() -> None:
    class _FakeRepository:
        def search_file_ids_by_metadata_query(
            self,
            *,
            user_id: int,
            query_text: str,
            file_ids: list[int] | None = None,
            limit: int = 20,
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, query_text, file_ids, limit, include_shared
            return list(range(1, 21))

    service = object.__new__(RetrievalPipelineService)
    service.repository = _FakeRepository()

    matches = service._metadata_prefilter_file_ids(
        question="region metropolitana de santiago",
        user_id=7,
        file_ids=None,
        limit=20,
    )

    assert matches == list(range(1, 21))


def test_retrieval_explicit_archive_scope_is_not_narrowed_by_metadata_prefilter() -> None:
    class _FakeEmbeddingService:
        def embed_query_text(self, *, text: str) -> list[float]:
            del text
            return [0.0, 0.1]

    class _FakeRepository:
        def search_file_ids_by_metadata_query(
            self,
            *,
            user_id: int,
            query_text: str,
            file_ids: list[int] | None = None,
            limit: int = 20,
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, query_text, file_ids, limit, include_shared
            return [201]

        def list_file_ids_for_input_filenames(
            self,
            *,
            user_id: int,
            file_names: list[str],
            file_ids: list[int] | None = None,
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, file_names, file_ids, include_shared
            return []

        def list_file_ids_for_archive_slugs(
            self,
            *,
            user_id: int,
            archive_slugs: list[str],
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, archive_slugs, include_shared
            return [201, 202, 203, 204, 205, 206]

        def get_archive_slug_map_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> dict[int, str]:
            del user_id, include_shared
            return {int(file_id): "LA122_ID_3979" for file_id in file_ids}

        def search_lexical_pages(
            self,
            *,
            user_id: int,
            question: str,
            file_ids: list[int] | None = None,
            limit: int = 20,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, question, file_ids, limit, include_shared
            return []

        def list_embeddings(
            self,
            *,
            file_ids: list[int],
            user_id: int,
            include_vectors: bool = False,
            modalities: list[str] | None = None,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del file_ids, user_id, include_vectors, modalities, include_shared
            return []

    class _FakeVectorStore:
        def similarity_search(
            self,
            *,
            query_vector: list[float],
            user_id: int | None = None,
            file_ids: list[int] | None = None,
            modality: str | None = None,
            top_k: int = 5,
            include_shared: bool = False,
        ) -> list[OracleVectorSearchResult]:
            del query_vector, user_id, top_k, include_shared
            if modality != "ocr_text":
                return []
            return [
                OracleVectorSearchResult(
                    file_id=int(file_id),
                    file_name=f"LA122_doc_{file_id}.pdf",
                    archive_slug="LA122_ID_3979",
                    file_code="LA122",
                    page_id=int(file_id) * 10,
                    page_number=1,
                    score=0.9,
                    summary_text=f"Evidencia del documento {file_id}.",
                    image_path_local="",
                    object_name_page="",
                    modality="ocr_text",
                    extraction_method="docling_rapidocr",
                    ocr_confidence=0.95,
                )
                for file_id in list(file_ids or [])
            ]

    service = object.__new__(RetrievalPipelineService)
    service.embedding_service = _FakeEmbeddingService()
    service.repository = _FakeRepository()
    service.oracle_vector_store = _FakeVectorStore()
    service.rerank_service = object()
    service.settings = Settings(_env_file=None)

    result = service.retrieve(
        question="/file:LA122_ID_3979 Que documentos integran el expediente?",
        user_id=7,
        file_ids=[201, 202, 203, 204, 205, 206],
        archive_slugs=["LA122_ID_3979"],
        top_k=6,
        candidate_k=24,
        min_pages_per_selected_doc=1,
        summary_mode="per_document",
        question_class="exhaustive_synthesis",
        scope_origin="metadata",
    )

    assert result.telemetry["metadata_prefilter_applied"] is False
    assert result.telemetry["doc_shortlist_count"] == 6
    assert {item.file_id for item in result.evidence} == {201, 202, 203, 204, 205, 206}


def test_retrieval_explicit_pdf_full_document_request_keeps_all_pages_in_order() -> None:
    class _FakeEmbeddingService:
        def embed_query_text(self, *, text: str) -> list[float]:
            del text
            return [0.0, 0.1]

    class _FakeRepository:
        def search_file_ids_by_metadata_query(
            self,
            *,
            user_id: int,
            query_text: str,
            file_ids: list[int] | None = None,
            limit: int = 20,
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, query_text, file_ids, limit, include_shared
            return []

        def list_file_ids_for_input_filenames(
            self,
            *,
            user_id: int,
            file_names: list[str],
            file_ids: list[int] | None = None,
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, include_shared
            if "AI041.pdf" not in file_names:
                return []
            allowed = {int(file_id) for file_id in list(file_ids or []) if int(file_id) > 0}
            return [701] if not allowed or 701 in allowed else []

        def list_file_ids_for_archive_slugs(
            self,
            *,
            user_id: int,
            archive_slugs: list[str],
            include_shared: bool = False,
        ) -> list[int]:
            del user_id, archive_slugs, include_shared
            return []

        def get_archive_slug_map_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> dict[int, str]:
            del user_id, include_shared
            return {int(file_id): "AI041_ID_49" for file_id in file_ids}

        def search_lexical_pages(
            self,
            *,
            user_id: int,
            question: str,
            file_ids: list[int] | None = None,
            limit: int = 20,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, question, file_ids, limit, include_shared
            return []

        def list_embeddings(
            self,
            *,
            file_ids: list[int],
            user_id: int,
            include_vectors: bool = False,
            modalities: list[str] | None = None,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, include_vectors, modalities, include_shared
            if 701 not in {int(file_id) for file_id in file_ids}:
                return []
            return [
                {
                    "file_id": 701,
                    "file_pages_id": 9000 + page_number,
                    "file_pages_number": page_number,
                    "file_pages_image_path_local": "",
                    "file_pages_output_obj_name": "",
                    "file_pages_ocr_confidence": 0.96,
                    "file_pages_ocr_method": "docling_rapidocr",
                    "file_pages_ocr_text": f"Texto OCR pagina {page_number}.",
                    "file_pages_visual_summary": "",
                    "file_pages_search_text": f"Texto OCR pagina {page_number}.",
                    "file_input_file_name": "AI041.pdf",
                    "archive_slug": "AI041_ID_49",
                    "file_code": "AI041",
                }
                for page_number in range(1, 7)
            ]

    class _FakeVectorStore:
        def similarity_search(
            self,
            *,
            query_vector: list[float],
            user_id: int | None = None,
            file_ids: list[int] | None = None,
            modality: str | None = None,
            top_k: int = 5,
            include_shared: bool = False,
        ) -> list[OracleVectorSearchResult]:
            del query_vector, user_id, top_k, include_shared
            if modality != "ocr_text" or 701 not in list(file_ids or []):
                return []
            return [
                OracleVectorSearchResult(
                    file_id=701,
                    file_name="AI041.pdf",
                    archive_slug="AI041_ID_49",
                    file_code="AI041",
                    page_id=9003,
                    page_number=3,
                    score=0.99,
                    summary_text="Texto OCR pagina 3.",
                    image_path_local="",
                    object_name_page="",
                    modality="ocr_text",
                    extraction_method="docling_rapidocr",
                    ocr_confidence=0.96,
                )
            ]

    service = object.__new__(RetrievalPipelineService)
    service.embedding_service = _FakeEmbeddingService()
    service.repository = _FakeRepository()
    service.oracle_vector_store = _FakeVectorStore()
    service.rerank_service = object()
    service.settings = Settings(_env_file=None)

    question = "Analiza todo el documento AI041.pdf y muestrame una lista completa clave valor."
    result = service.retrieve(
        question=question,
        user_id=7,
        file_ids=[701],
        top_k=5,
        question_class="exhaustive_synthesis",
        scope_origin="manual",
    )

    assert question_requests_full_document_coverage(question) is True
    assert result.telemetry["explicit_file_scope_applied"] is True
    assert result.telemetry["full_document_coverage_requested"] is True
    assert [item.page_number for item in result.evidence] == [1, 2, 3, 4, 5, 6]


def test_fact_resolver_expands_archive_scope_with_shared_documents() -> None:
    class _SharedArchiveMapRepository:
        def __init__(self) -> None:
            self.include_shared_seen: bool | None = None

        def get_archive_slug_map_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> dict[int, str]:
            del user_id
            self.include_shared_seen = include_shared
            return {int(file_id): "LA122_ID_3979" for file_id in file_ids}

    file_repository = _SharedArchiveMapRepository()
    resolver = QuestionFactResolver(repository=object(), file_repository=file_repository)

    expanded = resolver._expand_document_evidence_file_ids(
        user_id=7,
        candidate_file_ids=[201, 202, 203, 204, 205, 206],
        metadata_rows=[
            ArchiveMetadataEntry(
                file_id=201,
                archive_slug="LA122_ID_3979",
                fields={"Estado Contrato": "Vigente"},
            )
        ],
    )

    assert file_repository.include_shared_seen is True
    assert expanded == [201, 202, 203, 204, 205, 206]


def test_representative_questions_request_extra_page_coverage() -> None:
    assert question_requests_representative_details(
        "Que personas o representantes aparecen con facultades para firmar?"
    )
    assert not question_requests_representative_details("Que documentos integran el expediente?")


def test_per_doc_fallback_keeps_multiple_pages_per_document_for_quota() -> None:
    class _FakeRepository:
        def list_embeddings(
            self,
            *,
            file_ids: list[int],
            user_id: int,
            include_vectors: bool = False,
            modalities: list[str] | None = None,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, include_vectors, modalities, include_shared
            rows: list[dict[str, object]] = []
            for file_id in file_ids:
                rows.extend(
                    [
                        {
                            "file_id": file_id,
                            "file_input_file_name": f"doc-{file_id}.pdf",
                            "file_pages_id": int(file_id) * 10 + 1,
                            "file_pages_number": 1,
                            "file_pages_ocr_text": "comparecen representantes con facultades para firmar",
                            "file_pages_visual_summary": "",
                            "file_pages_image_path_local": "",
                            "file_pages_output_obj_name": "",
                            "file_pages_ocr_method": "docling_rapidocr",
                            "file_pages_ocr_confidence": 0.95,
                        },
                        {
                            "file_id": file_id,
                            "file_input_file_name": f"doc-{file_id}.pdf",
                            "file_pages_id": int(file_id) * 10 + 2,
                            "file_pages_number": 2,
                            "file_pages_ocr_text": "representada por segunda persona y mandato especial",
                            "file_pages_visual_summary": "",
                            "file_pages_image_path_local": "",
                            "file_pages_output_obj_name": "",
                            "file_pages_ocr_method": "docling_rapidocr",
                            "file_pages_ocr_confidence": 0.95,
                        },
                    ]
                )
            return rows

    service = object.__new__(RetrievalPipelineService)
    service.repository = _FakeRepository()

    candidates = service._build_per_doc_fallback_candidates(
        question="Que personas o representantes aparecen con facultades para firmar?",
        user_id=7,
        selected_file_ids=[501],
    )

    assert [item.evidence.page_number for item in candidates if item.evidence.file_id == 501] == [1, 2]


def test_representative_excerpt_stitches_people_names_across_pages() -> None:
    items = GraphSynthesis._build_per_document_items(
        [
            EvidenceItem(
                source_number=1,
                file_id=501,
                file_name="LA122_Modificacion.pdf",
                archive_slug="LA122_ID_3979",
                page_id=5011,
                page_number=1,
                score=0.95,
                summary_text=(
                    "comparecen SOCIEDAD TRANSPORTES COSTANERA S.A., representada por "
                    "don MARIO CARLOS PACHECO VAZQUEZ y por dona JANETTE LUCILA MANSILLA"
                ),
                image_path_local="",
            ),
            EvidenceItem(
                source_number=2,
                file_id=501,
                file_name="LA122_Modificacion.pdf",
                archive_slug="LA122_ID_3979",
                page_id=5012,
                page_number=2,
                score=0.60,
                summary_text=(
                    "TOLEDO, chilena, y por la otra ENTEL PCS TELECOMUNICACIONES S.A., "
                    "representada por don FRANCISCO JAVIER SPRENGER ARROYO."
                ),
                image_path_local="",
            ),
        ],
        question="Que personas o representantes aparecen con facultades para firmar?",
    )

    assert len(items) == 1
    excerpt = str(items[0]["summary_excerpt"])
    assert "JANETTE LUCILA MANSILLA" in excerpt
    assert "TOLEDO" in excerpt
    assert "FRANCISCO JAVIER SPRENGER ARROYO" in excerpt


def test_retrieval_builds_adjacent_pages_for_representative_boundary_context() -> None:
    class _FakeRepository:
        def list_embeddings(
            self,
            *,
            file_ids: list[int],
            user_id: int,
            include_vectors: bool = False,
            modalities: list[str] | None = None,
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, include_vectors, modalities, include_shared
            rows: list[dict[str, object]] = []
            for file_id in file_ids:
                rows.extend(
                    [
                        {
                            "file_id": file_id,
                            "file_input_file_name": "LA122_Modificacion.pdf",
                            "archive_slug": "LA122_ID_3979",
                            "file_pages_id": int(file_id) * 10 + 1,
                            "file_pages_number": 1,
                            "file_pages_ocr_text": "comparecen JANETTE LUCILA MANSILLA",
                            "file_pages_visual_summary": "",
                            "file_pages_image_path_local": "",
                            "file_pages_output_obj_name": "",
                            "file_pages_ocr_method": "docling_rapidocr",
                            "file_pages_ocr_confidence": 0.95,
                        },
                        {
                            "file_id": file_id,
                            "file_input_file_name": "LA122_Modificacion.pdf",
                            "archive_slug": "LA122_ID_3979",
                            "file_pages_id": int(file_id) * 10 + 2,
                            "file_pages_number": 2,
                            "file_pages_ocr_text": "TOLEDO, chilena, representada por FRANCISCO JAVIER SPRENGER ARROYO",
                            "file_pages_visual_summary": "",
                            "file_pages_image_path_local": "",
                            "file_pages_output_obj_name": "",
                            "file_pages_ocr_method": "docling_rapidocr",
                            "file_pages_ocr_confidence": 0.95,
                        },
                    ]
                )
            return rows

    service = object.__new__(RetrievalPipelineService)
    service.repository = _FakeRepository()
    selected = [
        EvidenceItem(
            source_number=1,
            file_id=501,
            file_name="LA122_Modificacion.pdf",
            archive_slug="LA122_ID_3979",
            page_id=5011,
            page_number=1,
            score=0.95,
            summary_text="comparecen JANETTE LUCILA MANSILLA",
            image_path_local="",
        )
    ]

    candidates = service._build_adjacent_page_candidates(
        user_id=7,
        selected_evidence=selected,
        selected_file_ids=[501],
    )

    assert [item.evidence.page_number for item in candidates] == [2]
    assert "TOLEDO" in candidates[0].evidence.summary_text


def test_final_evidence_quota_preserves_multiple_pages_per_document_when_trimming() -> None:
    evidence = [
        EvidenceItem(
            source_number=index,
            file_id=file_id,
            file_name=f"doc-{file_id}.pdf",
            archive_slug="LA122_ID_3979",
            page_id=file_id * 10 + page,
            page_number=page,
            score=1.0 / index,
            summary_text=f"doc {file_id} page {page}",
            image_path_local="",
            extraction_method="adjacent_page_context" if page == 2 else "",
        )
        for index, (file_id, page) in enumerate(
            [
                (501, 1),
                (502, 1),
                (503, 1),
                (501, 3),
                (502, 3),
                (503, 3),
                (501, 2),
                (502, 2),
                (503, 2),
            ],
            start=1,
        )
    ]

    limited = RetrievalPipelineService._enforce_final_evidence_quota(
        reranked=evidence[:6],
        candidate_pool=evidence,
        selected_file_ids=[501, 502, 503],
        min_pages_per_doc=2,
        desired_final=6,
    )

    pages_by_doc = {
        file_id: [item.page_number for item in limited if item.file_id == file_id]
        for file_id in [501, 502, 503]
    }
    assert pages_by_doc == {501: [1, 2], 502: [1, 2], 503: [1, 2]}


def test_archive_metadata_context_prioritizes_dynamic_fields_before_context_cutoff() -> None:
    class _FakeRepository:
        def get_archive_metadata_for_file_ids(
            self,
            *,
            user_id: int,
            file_ids: list[int],
            include_shared: bool = False,
        ) -> list[dict[str, object]]:
            del user_id, file_ids, include_shared
            fields = {f"Campo Flexible {index}": f"valor largo {index}" for index in range(80)}
            fields.update(
                {
                    "Renta o Precio Vigente": 504,
                    "Tipo de Moneda": "UF",
                    "Fecha de Término del Contrato": "01/08/2027",
                    "Estado Contrato": "Terminado",
                    "Estado Actividad": "Inactivo",
                }
            )
            return [
                {
                    "archive_slug": "LA122_ID_3979",
                    "metadata_json": json.dumps({"file": "LA122_ID_3979", "fields": fields}, ensure_ascii=False),
                }
            ]

    resolver = QuestionFactResolver(repository=object(), file_repository=_FakeRepository())

    context = resolver._build_archive_metadata_context(user_id=7, file_ids=[501])

    assert "Renta o Precio Vigente=504" in context
    assert "Tipo de Moneda=UF" in context
    assert "Fecha de Término del Contrato=01/08/2027" in context
    assert "Estado Contrato=Terminado" in context
    assert "Estado Actividad=Inactivo" in context


def test_fact_context_summary_prioritizes_dynamic_metadata_values_before_truncation() -> None:
    filler = "; ".join(f"Campo Flexible {index}=valor largo {index}" for index in range(80))
    fact_context = (
        "Archive metadata context:\n"
        "LA122_ID_3979: "
        f"{filler}; "
        "Renta o Precio Vigente=504; Tipo de Moneda=UF; "
        "Fecha de Inicio de Vigencia del Contrato=01/08/2025; "
        "Fecha de Término del Contrato=01/08/2027; "
        "Estado Contrato=Terminado; Estado Actividad=Inactivo\n"
        "Document inventory context:\n"
        "- file_id=501 archive=LA122_ID_3979 file=LA122_Modificacion.pdf status=completed pages=8"
    )

    summary = GraphSynthesis._extract_fact_context_summary(fact_context)

    assert "Renta o Precio Vigente=504" in summary
    assert "Tipo de Moneda=UF" in summary
    assert "Fecha de Inicio de Vigencia del Contrato=01/08/2025" in summary
    assert "Fecha de Término del Contrato=01/08/2027" in summary
    assert "Estado Contrato=Terminado" in summary
    assert "Estado Actividad=Inactivo" in summary


def test_build_oracle_text_contains_query_sanitizes_metadata_question() -> None:
    query = build_oracle_text_contains_query(
        "Usa la metadata para ubicar RM797_ID_1668.zip con Id 1668; "
        "luego confirma el Nombre de Propietario Principal y la Dirección."
    )

    tokens = query.split(" OR ")

    assert "{rm797_id_1668}" in tokens
    assert "{1668}" in tokens
    assert "{propietario}" in tokens
    assert "{dirección}" in tokens
    assert ";" not in query
    assert "." not in query
    assert ":" not in query


def test_build_oracle_text_contains_query_wraps_reserved_terms_as_literals() -> None:
    query = build_oracle_text_contains_query(
        "Busca RM797_ID_1668 and within about NOT owner"
    )

    tokens = query.split(" OR ")

    assert "{rm797_id_1668}" in tokens
    assert "{within}" in tokens
    assert "{about}" in tokens
    assert "{owner}" in tokens
    assert " within " not in query


def test_build_oracle_text_contains_query_returns_empty_for_only_noise() -> None:
    assert build_oracle_text_contains_query("de, la; y. a o") == ""


def test_rag_bootstrap_sql_aligns_archive_slug_and_oracle_text_defaults() -> None:
    files_sql = Path("apps/backend/db/bootstrap/sql/05_files.sql").read_text(encoding="utf-8")
    file_pages_sql = Path("apps/backend/db/bootstrap/sql/06_file_pages.sql").read_text(encoding="utf-8")
    page_embeddings_sql = Path("apps/backend/db/bootstrap/sql/07_page_embeddings.sql").read_text(encoding="utf-8")
    file_embeddings_sql = Path("apps/backend/db/bootstrap/sql/08_file_embeddings.sql").read_text(encoding="utf-8")
    uploads_sql = Path("apps/backend/db/bootstrap/sql/22_archive_metadata_uploads.sql").read_text(encoding="utf-8")
    upload_rows_sql = Path("apps/backend/db/bootstrap/sql/23_archive_metadata_upload_rows.sql").read_text(encoding="utf-8")
    archive_metadata_sql = Path("apps/backend/db/bootstrap/sql/24_archive_metadata.sql").read_text(encoding="utf-8")

    normalized_files_sql = _normalize_sql_whitespace(files_sql.lower())
    normalized_file_pages_sql = _normalize_sql_whitespace(file_pages_sql.lower())
    normalized_page_embeddings_sql = _normalize_sql_whitespace(page_embeddings_sql.lower())
    normalized_file_embeddings_sql = _normalize_sql_whitespace(file_embeddings_sql.lower())
    normalized_uploads_sql = _normalize_sql_whitespace(uploads_sql.lower())
    normalized_upload_rows_sql = _normalize_sql_whitespace(upload_rows_sql.lower())
    normalized_archive_metadata_sql = _normalize_sql_whitespace(archive_metadata_sql.lower())

    assert "archive_slug varchar2(256)" in normalized_files_sql
    assert "create index idx_files_user_archive_slug on files (user_id, archive_slug);" in normalized_files_sql
    assert "file_type_key" not in normalized_files_sql
    assert "parameters ('maintenance auto')" in normalized_file_pages_sql
    assert "archive_slug varchar2(256)" in normalized_page_embeddings_sql
    assert "create index idx_page_embeddings_archive on page_embeddings (user_id, archive_slug, page_embeddings_modality);" in normalized_page_embeddings_sql
    assert "include (user_id, file_id, file_pages_id, archive_slug, page_embeddings_modality)" in normalized_page_embeddings_sql
    assert "create vector index idx_page_embeddings_vector" in normalized_page_embeddings_sql
    assert "organization neighbor partitions" in normalized_page_embeddings_sql
    assert "organization inmemory neighbor graph" not in normalized_page_embeddings_sql
    assert "archive_slug varchar2(256)" in normalized_file_embeddings_sql
    assert "create index idx_file_embeddings_archive on file_embeddings (user_id, archive_slug);" in normalized_file_embeddings_sql
    assert "parameters ('maintenance auto')" in normalized_file_embeddings_sql
    assert "create vector index idx_file_embeddings_vector" in normalized_file_embeddings_sql
    assert "organization neighbor partitions" in normalized_file_embeddings_sql
    assert "organization inmemory neighbor graph" not in normalized_file_embeddings_sql
    assert "include (user_id, file_id, archive_slug)" in normalized_file_embeddings_sql
    assert "create table archive_metadata_uploads" in normalized_uploads_sql
    assert "display_name varchar2(300)" in normalized_uploads_sql
    assert "metadata_status varchar2(32) default 'active' not null" in normalized_uploads_sql
    assert "check (metadata_status in ('active', 'archived'))" in normalized_uploads_sql
    assert "create index idx_archive_metadata_uploads_status" in normalized_uploads_sql
    assert "create table archive_metadata_upload_rows" in normalized_upload_rows_sql
    assert "create index idx_archive_metadata_upload_rows_text on archive_metadata_upload_rows (search_text) indextype is ctxsys.context parameters ('maintenance auto');" in normalized_upload_rows_sql
    assert "create table archive_metadata" in normalized_archive_metadata_sql
    assert "create index idx_archive_metadata_text on archive_metadata (metadata_search_text) indextype is ctxsys.context parameters ('maintenance auto');" in normalized_archive_metadata_sql


def test_metadata_workbook_normalization_derives_file_column_from_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "info-docs.xlsx"
    output_path = tmp_path / "info-docs.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Prueba"
    worksheet.append(["Id", "Codigo de Sitio", "Nombre de Sitio", "Activo"])
    worksheet.append([49, "AI041", "Antena 41", True])
    worksheet.append([5515, "RM797", "Contrato Norte", False])
    workbook.save(source_path)

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.sheet_name == "Prueba"
    assert result.columns == ["file", "Id", "Codigo de Sitio", "Nombre de Sitio", "Activo"]
    assert result.total_rows == 2
    assert result.derived_file_column is True

    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Codigo de Sitio": "AI041",
            "Nombre de Sitio": "Antena 41",
            "Activo": "true",
        },
        {
            "file": "RM797_ID_5515",
            "Id": "5515",
            "Codigo de Sitio": "RM797",
            "Nombre de Sitio": "Contrato Norte",
            "Activo": "false",
        },
    ]


def test_metadata_workbook_normalization_reorders_existing_file_column(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_path = tmp_path / "metadata.xls"
    output_path = tmp_path / "metadata.csv"

    monkeypatch.setitem(
        normalize_metadata_workbook_to_csv.__globals__,
        "_read_xls_sheet",
        lambda source_path, sheet_name=None: LoadedWorkbookSheet(
            sheet_name="Legacy",
            headers=["Id", "File", "Region"],
            rows=[
                {"Id": "49", "File": "AI041_ID_49.zip", "Region": "Norte"},
            ],
        ),
    )

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.columns == ["file", "Id", "Region"]
    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Region": "Norte",
        }
    ]


def test_metadata_workbook_normalization_rejects_duplicate_derived_files(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "duplicated.xlsx"
    output_path = tmp_path / "duplicated.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Id", "Codigo de Sitio"])
    worksheet.append([49, "AI041"])
    worksheet.append([49, "AI041"])
    workbook.save(source_path)

    with pytest.raises(MetadataWorkbookNormalizationError, match="Duplicate `file` values"):
        normalize_metadata_workbook_to_csv(
            source_path=source_path,
            output_path=output_path,
        )


def test_metadata_workbook_normalization_disambiguates_duplicate_headers(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "duplicate_headers.xlsx"
    output_path = tmp_path / "duplicate_headers.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Id", "Codigo de Sitio", "Región", "Región", "Dirección", "Dirección"])
    worksheet.append([49, "AI041", "Norte", "Norte 2", "Uno", "Dos"])
    workbook.save(source_path)

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.columns == [
        "file",
        "Id",
        "Codigo de Sitio",
        "Región",
        "Región.1",
        "Dirección",
        "Dirección.1",
    ]
    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Codigo de Sitio": "AI041",
            "Región": "Norte",
            "Región.1": "Norte 2",
            "Dirección": "Uno",
            "Dirección.1": "Dos",
        }
    ]


def test_nomic_provider_uses_document_and_query_prefixes() -> None:
    provider = object.__new__(NomicLocalMultimodalProvider)
    calls: list[tuple[str, str]] = []

    def _fake_embed_prefixed_text(self, *, text: str, prefix: str) -> list[float]:
        calls.append((prefix, text))
        return [float(len(calls))]

    provider._embed_prefixed_text = MethodType(_fake_embed_prefixed_text, provider)

    assert provider.embed_document_text(text="documento") == [1.0]
    assert provider.embed_query_text(text="consulta") == [2.0]
    assert calls == [
        ("search_document", "documento"),
        ("search_query", "consulta"),
    ]


def test_embedding_service_routes_document_and_query_embeddings(monkeypatch: pytest.MonkeyPatch) -> None:
    class _FakeProvider:
        def __init__(self) -> None:
            self.calls: list[tuple[str, str]] = []

        def embed_document_text(self, *, text: str) -> list[float]:
            self.calls.append(("document", text))
            return [1.0]

        def embed_query_text(self, *, text: str) -> list[float]:
            self.calls.append(("query", text))
            return [2.0]

        def embed_image(self, *, image_path: Path, context_text: str = "") -> tuple[list[float], str]:
            raise AssertionError("image embeddings are not part of this unit test")

    fake_provider = _FakeProvider()
    monkeypatch.setattr(
        "apps.backend.app.rag.embedding_service.get_nomic_local_provider",
        lambda: fake_provider,
    )
    service = EmbeddingService(_build_settings())

    assert service.embed_document_text(text="archivo") == [1.0]
    assert service.embed_query_text(text="pregunta") == [2.0]
    assert service.embed_text(text="otra pregunta", input_type="query") == [2.0]
    assert fake_provider.calls == [
        ("document", "archivo"),
        ("query", "pregunta"),
        ("query", "otra pregunta"),
    ]


def test_rebalance_file_ids_by_archive_scope_round_robins_archives() -> None:
    ranked_file_ids = [664, 665, 666, 668, 669, 670]
    file_archive_map = {
        664: "RM797_ID_1668",
        665: "RM797_ID_1668",
        666: "RM797_ID_1668",
        668: "RM797_ID_5515",
        669: "RM797_ID_5515",
        670: "RM797_ID_5515",
    }

    balanced = RetrievalPipelineService._rebalance_file_ids_by_archive_scope(
        ranked_file_ids=ranked_file_ids,
        file_archive_map=file_archive_map,
        preferred_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
        limit=6,
    )

    assert balanced[:4] == [664, 668, 665, 669]


def test_prioritize_final_evidence_by_archive_scope_preserves_cross_archive_coverage() -> None:
    def _evidence(*, file_id: int, page_id: int, file_name: str, page_number: int) -> EvidenceItem:
        return EvidenceItem(
            source_number=0,
            file_id=file_id,
            file_name=file_name,
            file_code=None,
            page_id=page_id,
            page_number=page_number,
            score=0.9,
            summary_text=f"summary-{file_id}-{page_id}",
            image_path_local="",
            object_name_page="",
            extraction_method="ocr_text",
            ocr_confidence=0.99,
        )

    reranked = [
        _evidence(file_id=664, page_id=1, file_name="RM797_Contrato.pdf", page_number=1),
        _evidence(file_id=665, page_id=2, file_name="RM797_-_Decreto_MOP_Exento_N667.pdf", page_number=1),
        _evidence(file_id=666, page_id=3, file_name="RM797_-_Decreto_MOP_Exento_N668.pdf", page_number=1),
    ]
    candidate_pool = reranked + [
        _evidence(file_id=668, page_id=4, file_name="RM797_-_Contrato_2.pdf", page_number=1),
    ]
    file_archive_map = {
        664: "RM797_ID_1668",
        665: "RM797_ID_1668",
        666: "RM797_ID_1668",
        668: "RM797_ID_5515",
    }

    limited = RetrievalPipelineService._prioritize_final_evidence_by_archive_scope(
        reranked=reranked,
        candidate_pool=candidate_pool,
        preferred_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
        file_archive_map=file_archive_map,
        min_pages_per_archive=1,
        desired_final=3,
    )

    assert len(limited) == 3
    leading_archives = {
        RetrievalPipelineService._normalize_archive_slug(file_archive_map[item.file_id])
        for item in limited[:2]
    }
    assert leading_archives == {"rm797_id_1668", "rm797_id_5515"}


def test_prioritize_candidate_pool_by_archive_scope_promotes_best_candidate_per_archive() -> None:
    def _candidate(*, file_id: int, page_id: int, file_name: str, score: float) -> object:
        evidence = EvidenceItem(
            source_number=0,
            file_id=file_id,
            file_name=file_name,
            file_code=None,
            page_id=page_id,
            page_number=1,
            score=score,
            summary_text=f"summary-{file_id}-{page_id}",
            image_path_local="",
            object_name_page="",
            extraction_method="ocr_text",
            ocr_confidence=0.99,
        )
        return type(
            "CandidateStub",
            (),
            {
                "evidence": evidence,
                "fused_score": score,
            },
        )()

    candidates = [
        _candidate(file_id=664, page_id=1, file_name="RM797_Contrato.pdf", score=0.50),
        _candidate(file_id=668, page_id=4, file_name="RM797_-_Contrato_2.pdf", score=0.91),
        _candidate(file_id=665, page_id=2, file_name="RM797_-_Decreto_MOP_Exento_N667.pdf", score=0.95),
    ]
    file_archive_map = {
        664: "RM797_ID_1668",
        665: "RM797_ID_1668",
        668: "RM797_ID_5515",
    }

    ordered = RetrievalPipelineService._prioritize_candidate_pool_by_archive_scope(
        candidates=candidates,
        preferred_archive_slugs=["RM797_ID_1668", "RM797_ID_5515"],
        file_archive_map=file_archive_map,
        min_pages_per_archive=1,
    )

    assert [item.evidence.file_id for item in ordered[:2]] == [665, 668]
